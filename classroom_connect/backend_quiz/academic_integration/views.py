import logging
from typing import Any, Dict
import json
import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Avg
from django.http import HttpRequest, HttpResponse, JsonResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse
from django.utils import timezone

from .forms import (
    BatchEnrollmentForm, CourseForm, CSVUploadForm, StaffLoginForm,
    StudentAddForm, StudentForm, StudentLoginForm
)

# Sync functionality is in views_sync.py, imported directly in urls.py
# Import the API base URL function from utils
from .utils import api_base_url

# Define _api_base_url as an alias to api_base_url for backward compatibility
def _api_base_url():
    """Alias for api_base_url() function"""
    return api_base_url()

logger = logging.getLogger(__name__)


def _safe_json(response: requests.Response) -> Dict[str, Any]:
	"""
	Safely parse JSON from API response with enhanced error handling.
	"""
	try:
		result = response.json()
		# Log response status for debugging
		if not response.ok or not result.get("success", False):
			logger.warning(f"API request failed: Status {response.status_code}, "
						f"Response: {result.get('message', 'No message')}")
		return result
	except ValueError:
		logger.error(f"Failed to parse JSON from Academic Analyzer response (Status: {response.status_code}). "
					f"Content: {response.text[:200]}...", exc_info=True)
		return {"success": False, "message": "Invalid response format from API"}


def create_demo_quiz():
	"""
	Create a demo quiz if no quizzes exist in the database.
	This is a helper function for debugging purposes.
	"""
	from quiz.models import Quiz, Question, Choice, User
	from django.db.models import Q
	import logging
	import random
	
	logger = logging.getLogger(__name__)
	
	# Check if any quizzes exist
	if Quiz.objects.count() == 0:
		logger.info("No quizzes found. Creating a demo quiz.")
		
		# Find or create an admin user
		admin_user, created = User.objects.get_or_create(
			username="admin",
			defaults={
				"is_staff": True,
				"is_superuser": True,
				"email": "admin@example.com",
				"role": "admin"
			}
		)
		
		if created:
			admin_user.set_password("admin")
			admin_user.save()
		
		# Create a demo quiz
		demo_quiz = Quiz.objects.create(
			title="Demo Quiz",
			description="This is a demo quiz created automatically for testing.",
			is_active=True,
			course_id="DEMO101",
			created_by=admin_user,
			show_results=True,
			allow_review=True,
			quiz_type="tutorial",
			duration_minutes=15
		)
		
		# Add some questions
		for i in range(1, 4):
			question = Question.objects.create(
				quiz=demo_quiz,
				text=f"Demo Question {i}",
				question_type="mcq_single",
				order=i
			)
			
			# Add choices
			correct_choice = random.randint(1, 4)
			for j in range(1, 5):
				Choice.objects.create(
					question=question,
					text=f"Option {j}",
					is_correct=(j == correct_choice),
					order=j
				)
		
		logger.info(f"Created demo quiz with ID {demo_quiz.id}")
		return demo_quiz
	return None


def staff_login(request: HttpRequest) -> HttpResponse:
	if request.session.get("staff_email"):
		return redirect("academic_integration:staff_dashboard")

	form = StaffLoginForm(request.POST or None)

	if request.method == "POST" and form.is_valid():
		payload = form.cleaned_data
		try:
			response = requests.post(
				f"{api_base_url()}/staff/auth",
				json={"email": payload["email"], "password": payload["password"]},
				timeout=5,
			)
		except requests.RequestException:
			logger.exception("Staff auth request failed")
			form.add_error(None, "Cannot reach Academic Analyzer API right now. Please try again later.")
		else:
			body = _safe_json(response)
			if response.ok and body.get("success"):
				request.session["staff_email"] = body.get("email", payload["email"])
				request.session["staff_teacher_id"] = body.get("teacherId")
				request.session["staff_name"] = body.get("name") or body.get("email") or payload["email"]
				messages.success(request, "Logged in successfully.")
				return redirect("academic_integration:staff_dashboard")
			error_message = body.get("message", "Invalid credentials. Please try again.")
			form.add_error(None, error_message)

	return render(request, "academic_integration/staff_login.html", {"form": form})


def staff_dashboard(request: HttpRequest) -> HttpResponse:
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")

	courses: list[Dict[str, Any]] = []
	api_error: str | None = None

	try:
		logger.info(f"Loading dashboard data for staff: {staff_email}")
		response = requests.get(
			f"{api_base_url()}/staff/dashboard",
			params={"email": staff_email},
			timeout=10,  # Increased timeout for better reliability
		)
	except requests.RequestException as e:
		logger.exception(f"Failed to load staff dashboard data: {str(e)}")
		api_error = "Could not reach Academic Analyzer API. Please check your internet connection and refresh the page. If the issue persists, contact support."
	else:
		body = _safe_json(response)
		if response.ok and body.get("success"):
			courses = body.get("courses", [])
			if body.get("name"):
				request.session["staff_name"] = body["name"]
			logger.info(f"Successfully loaded dashboard with {len(courses)} courses")
		else:
			error_message = body.get("message", "Unknown error")
			logger.error(f"API Error in staff dashboard: {error_message}")
			api_error = f"API Error: {error_message}. Please try again later."

	context = {
		"staff_name": request.session.get("staff_name") or staff_email,
		"staff_email": staff_email,
		"courses": courses,
		"api_error": api_error,
	}
	return render(request, "academic_integration/staff_dashboard.html", context)


def admin_quiz_dashboard(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to manage quizzes - displays a list of quizzes created by the staff
	or for courses they teach.
	"""
	from django.db.models import Q, Avg
	from quiz.models import Quiz, QuizAttempt
	
	staff_email = request.session.get("staff_email")
	
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")
	
	# Get courses taught by the teacher
	courses = []
	try:
		response = requests.get(
			f"{api_base_url()}/staff/dashboard",
			params={"email": staff_email},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get('success'):
				courses = data.get('courses', [])
	except requests.RequestException:
		logger.exception("Failed to fetch courses for quiz dashboard")
	
	# Create a dictionary to store courses by ID
	course_dict = {course['courseId']: course for course in courses}
	
	# Include both course-specific quizzes and quizzes created by this staff
	course_ids = [course['courseId'] for course in courses]
	
	quizzes = Quiz.objects.filter(
		Q(course_id__in=course_ids) | 
		Q(created_by__email=staff_email) |
		Q(created_by__username=staff_email)
	).order_by('-created_at')
	
	# Get quiz statistics and enhance with course information
	for quiz in quizzes:
		quiz.num_attempts = QuizAttempt.objects.filter(quiz=quiz).count()
		quiz.num_completed = QuizAttempt.objects.filter(quiz=quiz, completed_at__isnull=False).count()
		quiz.avg_score = QuizAttempt.objects.filter(quiz=quiz, completed_at__isnull=False).aggregate(Avg('percentage'))['percentage__avg'] or 0
		
		# Check for submissions that need grading (status='submitted')
		quiz.needs_grading = QuizAttempt.objects.filter(
			quiz=quiz,
			status='submitted'
		).exists()
		
		# Add course information if available
		if quiz.course_id and quiz.course_id in course_dict:
			quiz.course_name = course_dict[quiz.course_id]['courseName']
		else:
			quiz.course_name = None
	
	context = {
		'quizzes': quizzes,
		'courses': courses,
		'staff_email': staff_email,
		'staff_name': request.session.get('staff_name', staff_email),
	}
	return render(request, "academic_integration/admin_quiz_dashboard.html", context)


def create_quiz(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to create a new quiz.
	"""
	from quiz.models import Quiz, Question, Choice, User
	import json
	
	# Ensure staff is logged in
	staff_email = request.session.get('staff_email')
	staff_id = request.session.get('staff_teacher_id')
	
	if not staff_email:
		messages.error(request, "You must be logged in as staff")
		return redirect('academic_integration:staff_login')
	
	if request.method == 'POST':
		try:
			data = json.loads(request.body)
			
			# Create or get the staff user
			staff_user, created = User.objects.get_or_create(
				username=staff_email,
				defaults={
					'email': staff_email,
					'role': 'admin'
				}
			)
			
			# Set is_mock_test based on tutorial_number
			quiz_type = data.get('quiz_type', 'tutorial')
			if not data.get('tutorial_number') and quiz_type == 'tutorial':
				quiz_type = 'mock'
			
			# Convert empty tutorial_number string to None
			tutorial_number = data.get('tutorial_number')
			if tutorial_number == '':
				tutorial_number = None
				
			# Create the quiz
			quiz = Quiz.objects.create(
				title=data['title'],
				description=data.get('description', ''),
				start_date=data.get('start_date'),
				complete_by_date=data.get('complete_by_date'),
				course_id=data.get('course_id'),
				tutorial_number=tutorial_number,
				created_by=staff_user,
				quiz_type=quiz_type,
				duration_minutes=int(data.get('duration_minutes', 30)),
				is_active=data.get('is_active', True),
				show_results=data.get('show_results', True),
				allow_review=data.get('allow_review', True)
			)
			
			# Create questions
			for question_data in data['questions']:
				question = Question.objects.create(
					quiz=quiz,
					text=question_data['text'],
					question_type=question_data['type'],
					points=question_data.get('points', 1),
					order=question_data.get('order', 0)
				)
				if question_data['type'] in ['mcq_single', 'mcq_multiple', 'true_false']:
					for choice_data in question_data['choices']:
						Choice.objects.create(
							question=question,
							text=choice_data['text'],
							is_correct=choice_data['is_correct'],
							order=choice_data.get('order', 0)
						)
			return JsonResponse({'success': True, 'quiz_id': quiz.id})
		except Exception as e:
			return JsonResponse({'success': False, 'error': str(e)})
	
	# Get courses for the dropdown menu
	courses = []
	try:
		response = requests.get(
			f"{api_base_url()}/staff/dashboard",
			params={"email": staff_email},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get('success'):
				courses = data.get('courses', [])
	except requests.RequestException:
		logger.exception("Failed to fetch courses for create quiz")
	
	context = {
		'courses': courses,
		'staff_email': staff_email,
		'staff_name': request.session.get('staff_name', staff_email),
	}
	return render(request, "academic_integration/create_quiz.html", context)


def edit_quiz(request: HttpRequest, quiz_id: int) -> HttpResponse:
	"""
	View for staff to edit an existing quiz.
	"""
	from quiz.models import Quiz, Question, Choice, User
	from django.shortcuts import get_object_or_404
	import json
	
	# Ensure staff is logged in
	staff_email = request.session.get('staff_email')
	staff_id = request.session.get('staff_teacher_id')
	
	if not staff_email:
		messages.error(request, "You must be logged in as staff")
		return redirect('academic_integration:staff_login')
	
	quiz = get_object_or_404(Quiz, pk=quiz_id)
	
	# Verify staff has access to this quiz
	handled_courses = []
	try:
		response = requests.get(
			f"{api_base_url()}/staff/dashboard",
			params={"email": staff_email},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get('success'):
				handled_courses = [course['courseId'] for course in data.get('courses', [])]
	except requests.RequestException:
		logger.exception("Failed to fetch courses for edit quiz")
	
	can_edit = False
	
	# Check if staff handles the course or created the quiz
	if quiz.course_id and quiz.course_id in handled_courses:
		can_edit = True
	elif quiz.created_by and (quiz.created_by.email == staff_email or quiz.created_by.username == staff_email):
		can_edit = True
	elif not quiz.course_id:  # Quizzes not linked to any course can be edited by any staff
		can_edit = True
	
	if not can_edit:
		messages.error(request, "You don't have permission to edit this quiz")
		return redirect('academic_integration:staff_dashboard')
	
	if request.method == 'POST':
		try:
			data = json.loads(request.body)
			quiz.title = data['title']
			quiz.description = data.get('description', '')
			quiz.start_date = data.get('start_date')
			quiz.complete_by_date = data.get('complete_by_date')
			quiz.course_id = data.get('course_id')
			
			# Convert empty tutorial_number string to None
			tutorial_number = data.get('tutorial_number')
			if tutorial_number == '':
				tutorial_number = None
			quiz.tutorial_number = tutorial_number
			
			quiz.quiz_type = data.get('quiz_type', 'tutorial')
			quiz.duration_minutes = int(data.get('duration_minutes', 30))
			quiz.is_active = data.get('is_active', True)
			quiz.show_results = data.get('show_results', True)
			quiz.allow_review = data.get('allow_review', True)
			
			# Create or get the staff user
			staff_user, created = User.objects.get_or_create(
				username=staff_email,
				defaults={
					'email': staff_email,
					'role': 'admin'
				}
			)
			
			# Set created_by if not already set
			if not quiz.created_by:
				quiz.created_by = staff_user
				
			quiz.save()
			
			# Delete existing questions
			quiz.questions.all().delete()
			
			# Create new questions
			for question_data in data['questions']:
				question = Question.objects.create(
					quiz=quiz,
					text=question_data['text'],
					question_type=question_data['type'],
					points=question_data.get('points', 1),
					order=question_data.get('order', 0)
				)
				if question_data['type'] in ['mcq_single', 'mcq_multiple', 'true_false']:
					for choice_data in question_data['choices']:
						Choice.objects.create(
							question=question,
							text=choice_data['text'],
							is_correct=choice_data['is_correct'],
							order=choice_data.get('order', 0)
						)
			
			return JsonResponse({'success': True})
		except Exception as e:
			return JsonResponse({'success': False, 'error': str(e)})
	
	# Get courses for the dropdown menu
	courses = []
	try:
		response = requests.get(
			f"{api_base_url()}/staff/dashboard",
			params={"email": staff_email},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get('success'):
				courses = data.get('courses', [])
	except requests.RequestException:
		logger.exception("Failed to fetch courses for edit quiz")
	
	context = {
		'quiz': quiz,
		'courses': courses,
		'staff_email': staff_email,
		'staff_name': request.session.get('staff_name', staff_email),
	}
	return render(request, "academic_integration/edit_quiz.html", context)


def delete_quiz(request: HttpRequest, quiz_id: int) -> HttpResponse:
    """
    View for staff to delete a quiz.
    """
    from quiz.models import Quiz
    from django.shortcuts import get_object_or_404
    
    # Ensure staff is logged in
    staff_email = request.session.get('staff_email')
    
    if not staff_email:
        messages.error(request, "You must be logged in as staff")
        return redirect('academic_integration:staff_login')
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    # Verify staff has access to this quiz
    handled_courses = []
    try:
        response = requests.get(
            f"{api_base_url()}/staff/dashboard",
            params={"email": staff_email},
            timeout=5,
        )
        if response.ok:
            data = _safe_json(response)
            if data.get('success'):
                handled_courses = [course['courseId'] for course in data.get('courses', [])]
    except requests.RequestException:
        logger.exception("Failed to fetch courses for delete quiz")
    
    can_delete = False
    
    # Check if staff handles the course or created the quiz
    if quiz.course_id and quiz.course_id in handled_courses:
        can_delete = True
    elif quiz.created_by and (quiz.created_by.email == staff_email or quiz.created_by.username == staff_email):
        can_delete = True
    
    if not can_delete:
        messages.error(request, "You don't have permission to delete this quiz")
        return redirect('academic_integration:staff_dashboard')
    
    if request.method == 'POST':
        quiz.delete()
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def get_quiz_data(request: HttpRequest, quiz_id: int) -> HttpResponse:
    """
    API endpoint to get quiz data for editing or taking.
    """
    from quiz.models import Quiz, QuizAttempt, User
    from academic_integration.models import Student
    from django.shortcuts import get_object_or_404
    
    # Check if staff or student is logged in
    staff_email = request.session.get('staff_email')
    student_roll_number = request.session.get('student_roll_number')
    
    if not staff_email and not student_roll_number:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    # Handle staff request
    if staff_email:
        # Verify staff has access to this quiz
        handled_courses = []
        try:
            response = requests.get(
                f"{api_base_url()}/staff/dashboard",
                params={"email": staff_email},
                timeout=5,
            )
            if response.ok:
                data = _safe_json(response)
                if data.get('success'):
                    handled_courses = [course['courseId'] for course in data.get('courses', [])]
        except requests.RequestException:
            logger.exception("Failed to fetch courses for quiz data")
        
        can_access = False
        
        # Check if staff handles the course or created the quiz
        if quiz.course_id and quiz.course_id in handled_courses:
            can_access = True
        elif quiz.created_by and (quiz.created_by.email == staff_email or quiz.created_by.username == staff_email):
            can_access = True
        elif not quiz.course_id:  # Quizzes not linked to any course can be accessed by any staff
            can_access = True
            
        if not can_access:
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Prepare quiz data
        questions_data = []
        for question in quiz.questions.all().order_by('order'):
            choices_data = []
            for choice in question.choices.all().order_by('order'):
                choices_data.append({
                    'text': choice.text,
                    'is_correct': choice.is_correct
                })
            
            questions_data.append({
                'id': question.id,
                'text': question.text,
                'question_type': question.question_type,
                'order': question.order,
                'points': question.points,
                'choices': choices_data
            })
        
        quiz_data = {
            'id': quiz.id,
            'title': quiz.title,
            'description': quiz.description,
            'quiz_type': quiz.quiz_type,
            'course_id': quiz.course_id,
            'tutorial_number': quiz.tutorial_number,
            'start_date': quiz.start_date.isoformat() if quiz.start_date else None,
            'complete_by_date': quiz.complete_by_date.isoformat() if quiz.complete_by_date else None,
            'duration_minutes': quiz.duration_minutes,
            'is_active': quiz.is_active,
            'show_results': quiz.show_results,
            'allow_review': quiz.allow_review,
            'questions': questions_data
        }
        
        return JsonResponse({'success': True, 'quiz': quiz_data})
    
    # Handle student request
    else:
        # Verify student is enrolled in the course
        if quiz.course_id:
            enrolled_courses = []
            try:
                response = requests.get(
                    f"{api_base_url()}/student/dashboard",
                    params={"rollno": student_roll_number},
                    timeout=5,
                )
                if response.ok:
                    data = _safe_json(response)
                    if data.get('success'):
                        enrolled_courses = [course['courseId'] for course in data.get('courses', [])]
            except requests.RequestException:
                logger.exception("Failed to fetch courses for quiz data")
                
            if quiz.course_id not in enrolled_courses:
                return JsonResponse({'success': False, 'error': 'You are not enrolled in this course'}, status=403)
        
        # Check if quiz is available using the quiz model's is_available property
        if not quiz.is_available:
            is_visible, reason = quiz.debug_visibility_status()
            return JsonResponse({'success': False, 'error': reason}, status=403)
            
        # Get student user and attempt
        student = get_object_or_404(Student, user__username=student_roll_number)
        
        # Check for existing attempts
        attempt = QuizAttempt.objects.filter(
            quiz=quiz,
            user=student.user
        ).order_by('-started_at').first()
        
        # Prepare quiz data for student
        questions_data = []
        for question in quiz.questions.all().order_by('order'):
            choices_data = []
            for choice in question.choices.all().order_by('order'):
                # Don't include is_correct flag for student
                choices_data.append({
                    'id': choice.id,
                    'text': choice.text
                })
                
            questions_data.append({
                'id': question.id,
                'text': question.text,
                'question_type': question.question_type,
                'choices': choices_data
            })
        
        quiz_data = {
            'id': quiz.id,
            'title': quiz.title,
            'description': quiz.description,
            'quiz_type': quiz.quiz_type,
            'duration_minutes': quiz.duration_minutes,
            'questions': questions_data
        }
        
        # Include attempt information
        attempt_data = None
        if attempt:
            # Calculate time remaining
            time_remaining_seconds = 0
            if attempt.started_at and not attempt.completed_at:
                elapsed_seconds = (timezone.now() - attempt.started_at).total_seconds()
                time_remaining_seconds = max(0, quiz.duration_minutes * 60 - elapsed_seconds)
            
            attempt_data = {
                'id': attempt.id,
                'started_at': attempt.started_at.isoformat() if attempt.started_at else None,
                'completed_at': attempt.completed_at.isoformat() if attempt.completed_at else None,
                'time_remaining_seconds': time_remaining_seconds
            }
        
        return JsonResponse({
            'success': True, 
            'quiz': quiz_data, 
            'attempt': attempt_data
        })


def end_quiz(request: HttpRequest, quiz_id: int) -> HttpResponse:
    """
    API endpoint to end a quiz.
    """
    from quiz.models import Quiz
    from django.shortcuts import get_object_or_404
    
    # Ensure staff is logged in
    staff_email = request.session.get('staff_email')
    
    if not staff_email:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    # Verify staff has access to this quiz
    handled_courses = []
    try:
        response = requests.get(
            f"{api_base_url()}/staff/dashboard",
            params={"email": staff_email},
            timeout=5,
        )
        if response.ok:
            data = _safe_json(response)
            if data.get('success'):
                handled_courses = [course['courseId'] for course in data.get('courses', [])]
    except requests.RequestException:
        logger.exception("Failed to fetch courses for end quiz")
    
    can_end = False
    
    # Check if staff handles the course or created the quiz
    if quiz.course_id and quiz.course_id in handled_courses:
        can_end = True
    elif quiz.created_by and (quiz.created_by.email == staff_email or quiz.created_by.username == staff_email):
        can_end = True
        
    if not can_end:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        quiz.is_ended = True
        quiz.save()
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)


def quiz_attempt(request: HttpRequest, quiz_id: int) -> HttpResponse:
    """
    API endpoint for students to start a quiz attempt.
    """
    from quiz.models import Quiz, QuizAttempt, User
    from academic_integration.models import Student
    from django.shortcuts import get_object_or_404
    import json
    
    # Ensure student is logged in
    student_roll_number = request.session.get('student_roll_number')
    student_id = request.session.get('student_id')
    
    if not student_roll_number:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    quiz = get_object_or_404(Quiz, pk=quiz_id, is_active=True)
    
    # Check if quiz is available using the quiz model's is_available property
    if not quiz.is_available:
        is_visible, reason = quiz.debug_visibility_status()
        return JsonResponse({'success': False, 'error': reason}, status=403)
    
    # Check if student is enrolled in the course
    if quiz.course_id:
        enrolled_courses = []
        try:
            response = requests.get(
                f"{api_base_url()}/student/dashboard",
                params={"rollno": student_roll_number},
                timeout=5,
            )
            if response.ok:
                data = _safe_json(response)
                if data.get('success'):
                    enrolled_courses = [course['courseId'] for course in data.get('courses', [])]
        except requests.RequestException:
            logger.exception("Failed to fetch courses for quiz attempt")
            
        if quiz.course_id not in enrolled_courses:
            return JsonResponse({'success': False, 'error': 'You are not enrolled in this course'}, status=403)
    
    # Get or create student user
    student_user, created = User.objects.get_or_create(
        username=student_roll_number,
        defaults={
            'email': f"{student_roll_number}@psgtech.ac.in",
            'role': 'student'
        }
    )
    
    # Get or create student profile
    student, created = Student.objects.get_or_create(
        user=student_user,
        defaults={
            'student_id': student_id or student_roll_number  # Use academic analyzer ID if available
        }
    )
    
    # Check for existing attempts
    attempt = QuizAttempt.objects.filter(
        quiz=quiz,
        user=student.user
    ).order_by('-started_at').first()
    
    # If there's already an attempt in progress, return it
    if attempt and attempt.started_at and not attempt.completed_at:
        # Calculate time remaining
        elapsed_seconds = (timezone.now() - attempt.started_at).total_seconds()
        time_remaining_seconds = max(0, quiz.duration_minutes * 60 - elapsed_seconds)
        
        return JsonResponse({
            'success': True,
            'attempt': {
                'id': attempt.id,
                'started_at': attempt.started_at.isoformat(),
                'time_remaining_seconds': time_remaining_seconds
            }
        })
    
    # If completed and no retakes allowed, deny attempt
    if attempt and attempt.completed_at and not quiz.allow_retake:
        return JsonResponse({'success': False, 'error': 'You have already completed this quiz'}, status=403)
    
    # Create a new attempt
    new_attempt = QuizAttempt.objects.create(
        quiz=quiz,
        user=student.user,
        started_at=timezone.now(),
        status='in_progress'
    )
    
    return JsonResponse({
        'success': True,
        'attempt': {
            'id': new_attempt.id,
            'started_at': new_attempt.started_at.isoformat(),
            'time_remaining_seconds': quiz.duration_minutes * 60
        }
    })


def submit_quiz(request: HttpRequest, quiz_id: int) -> HttpResponse:
    """
    API endpoint for students to submit quiz answers.
    """
    from quiz.models import Quiz, QuizAttempt, User, Question, Choice, QuizAnswer
    from academic_integration.models import Student
    from django.shortcuts import get_object_or_404
    import json
    import logging
    
    # Set up logging
    logger = logging.getLogger(__name__)
    logger.debug(f"Submit quiz request received for quiz_id: {quiz_id}")
    
    # Ensure student is logged in
    student_roll_number = request.session.get('student_roll_number')
    
    if not student_roll_number:
        logger.warning("Quiz submission attempted without authentication")
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    # Verify request method and content
    if request.method != 'POST':
        logger.warning(f"Invalid request method: {request.method} for quiz submission")
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    try:
        raw_body = request.body.decode('utf-8')
        logger.debug(f"Raw request body: {raw_body}")
        
        data = json.loads(request.body)
        answers = data.get('answers', {})
        logger.debug(f"Received answers data: {answers}")
        
        if not answers:
            logger.warning("No answers provided in submission")
            return JsonResponse({'success': False, 'error': 'No answers were provided. Please select at least one answer before submitting.'}, status=400)
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Invalid JSON data: {str(e)}'}, status=400)
    except Exception as e:
        logger.error(f"Unexpected error processing request body: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error processing request: {str(e)}'}, status=500)
        
    # Additional validation for answer format
    valid_answers = {}
    for question_key, answer_value in answers.items():
        try:
            # Extract question ID from key (format: "question_X")
            question_id = int(question_key.split('_')[1])
            valid_answers[question_key] = answer_value
            logger.debug(f"Validated answer for question {question_id}: {answer_value}")
        except (ValueError, IndexError) as e:
            logger.warning(f"Invalid question key format: {question_key}, error: {str(e)}")
            
    if not valid_answers:
        logger.warning("No valid answers after format validation")
        return JsonResponse({'success': False, 'error': 'No valid answers were provided.'}, status=400)
        
    answers = valid_answers
    
    # Get quiz and student
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    student = get_object_or_404(Student, user__username=student_roll_number)
    
    # Get the current attempt
    attempt = QuizAttempt.objects.filter(
        quiz=quiz,
        user=student.user,
        completed_at__isnull=True
    ).order_by('-started_at').first()
    
    if not attempt:
        # Check if already completed
        completed = QuizAttempt.objects.filter(
            quiz=quiz,
            user=student.user,
            completed_at__isnull=False
        ).exists()
        
        if completed:
            return JsonResponse({'success': False, 'error': 'You have already completed this quiz'})
        else:
            return JsonResponse({'success': False, 'error': 'No active quiz attempt found'})
    
    # Mark attempt as completed
    attempt.completed_at = timezone.now()
    attempt.duration_seconds = (attempt.completed_at - attempt.started_at).total_seconds()
    attempt.status = 'submitted'
    
    # Process answers and calculate score
    total_points = 0
    earned_points = 0
    
    logger.debug(f"Processing answers for {quiz.questions.count()} questions")
    
    for question in quiz.questions.all():
        answer_key = f"question_{question.id}"
        
        # Create a new answer record
        answer = QuizAnswer.objects.create(
            question=question,
            attempt=attempt
        )
        
        if answer_key in answers:
            answer_value = answers[answer_key]
            logger.debug(f"Processing answer for {answer_key}: {answer_value}")
            
            # Handle different question types
            if question.question_type == 'mcq_single':
                # Single choice question
                try:
                    # Convert answer_value to int if possible
                    if isinstance(answer_value, str) and answer_value.lower() == 'undefined':
                        logger.warning(f"Received 'undefined' as choice ID for question {question.id}")
                        # Skip this question - no valid answer provided
                        answer.is_correct = False
                        answer.points_earned = 0
                    else:
                        try:
                            # Try to convert to int if it's a string representing a number
                            if isinstance(answer_value, str) and answer_value.isdigit():
                                answer_value = int(answer_value)
                            elif isinstance(answer_value, str):
                                logger.warning(f"Non-numeric string answer value: {answer_value}")
                                answer.is_correct = False
                                answer.points_earned = 0
                                answer.save()
                                total_points += question.points
                                earned_points += answer.points_earned
                                continue
                        except (ValueError, TypeError) as e:
                            logger.warning(f"Failed to convert {answer_value} to int: {str(e)}")
                            answer.is_correct = False
                            answer.points_earned = 0
                            answer.save()
                            total_points += question.points
                            earned_points += answer.points_earned
                            continue
                            
                        try:
                            choice = Choice.objects.get(id=answer_value, question=question)
                            answer.selected_choices.add(choice)
                            logger.info(f"MCQ Single: Added choice {choice.id} ({choice.text}) for question {question.id}")
                            
                            if choice.is_correct:
                                answer.points_earned = question.points
                                answer.is_correct = True
                                logger.info(f"MCQ Single: Question {question.id} marked CORRECT - earned {question.points} points")
                            else:
                                answer.is_correct = False
                                answer.points_earned = 0
                                logger.info(f"MCQ Single: Question {question.id} marked INCORRECT - wrong choice selected")
                        except Choice.DoesNotExist:
                            # Invalid choice ID
                            logger.error(f"Choice with ID {answer_value} does not exist for question {question.id}")
                            answer.is_correct = False
                            answer.points_earned = 0
                except Exception as e:
                    logger.error(f"Error processing single choice answer: {str(e)}", exc_info=True)
                    # Don't award points if there was an error
                    answer.is_correct = False
                    answer.points_earned = 0
                    
            elif question.question_type == 'mcq_multiple':
                # Multiple choice question
                if isinstance(answer_value, list):
                    correct_choice_ids = set(question.choices.filter(is_correct=True).values_list('id', flat=True))
                    selected_choice_ids = set()
                    
                    logger.info(f"MCQ Multiple: Question {question.id} has {len(correct_choice_ids)} correct choices")
                    
                    # Add all selected choices
                    for choice_id in answer_value:
                        try:
                            # Handle 'undefined' choice IDs
                            if isinstance(choice_id, str) and choice_id.lower() == 'undefined':
                                logger.warning(f"Received 'undefined' as choice ID for multiple choice question {question.id}")
                                continue
                                
                            # Try to convert string to int if needed
                            if isinstance(choice_id, str) and choice_id.isdigit():
                                choice_id = int(choice_id)
                            elif isinstance(choice_id, str):
                                logger.warning(f"Non-numeric string choice ID: {choice_id}")
                                continue
                                
                            try:
                                choice = Choice.objects.get(id=choice_id, question=question)
                                answer.selected_choices.add(choice)
                                selected_choice_ids.add(choice.id)
                                logger.info(f"MCQ Multiple: Added choice {choice.id} ({choice.text})")
                            except Choice.DoesNotExist:
                                logger.error(f"Choice {choice_id} does not exist for question {question.id}")
                        except (ValueError, TypeError) as e:
                            logger.warning(f"Error processing choice {choice_id} for question {question.id}: {str(e)}")
                    
                    # Check if the selected choices exactly match the correct choices
                    if selected_choice_ids == correct_choice_ids and len(selected_choice_ids) > 0:
                        answer.points_earned = question.points
                        answer.is_correct = True
                        logger.info(f"MCQ Multiple: Question {question.id} marked CORRECT - all correct choices selected, no incorrect ones")
                    else:
                        answer.is_correct = False
                        answer.points_earned = 0
                        logger.info(f"MCQ Multiple: Question {question.id} marked INCORRECT - Selected: {selected_choice_ids}, Correct: {correct_choice_ids}")
                else:
                    # Single value provided for multiple choice - treat as array with one element
                    logger.warning(f"Single value {answer_value} provided for multiple choice question {question.id}")
                    try:
                        if isinstance(answer_value, str) and answer_value.isdigit():
                            answer_value = int(answer_value)
                        
                        choice = Choice.objects.get(id=answer_value, question=question)
                        answer.selected_choices.add(choice)
                        
                        # Check if this is the only correct choice
                        correct_choices = question.choices.filter(is_correct=True)
                        if correct_choices.count() == 1 and choice.is_correct:
                            answer.points_earned = question.points
                            answer.is_correct = True
                            logger.info(f"MCQ Multiple: Single choice {choice.id} is the only correct answer")
                        else:
                            answer.is_correct = False
                            answer.points_earned = 0
                            logger.info(f"MCQ Multiple: Single choice not sufficient or incorrect")
                    except (Choice.DoesNotExist, ValueError, TypeError) as e:
                        logger.error(f"Error processing single choice for MCQ multiple: {str(e)}")
                        answer.is_correct = False
                        answer.points_earned = 0
                        
            elif question.question_type == 'text':
                # Text question
                answer.text_answer = str(answer_value)
                
                # Check if answer matches exactly (case insensitive)
                if question.correct_answer is not None and answer.text_answer.lower() == question.correct_answer.lower():
                    answer.points_earned = question.points
                    answer.is_correct = True
                    
            elif question.question_type == 'true_false':
                # True/False question
                try:
                    logger.debug(f"Processing true/false answer: {answer_value} (type: {type(answer_value).__name__})")
                    
                    # For True/False questions, the answer is usually a choice ID
                    # We need to check if the selected choice is marked as correct
                    selected_choice = None
                    
                    # Convert the answer to find the selected choice
                    if isinstance(answer_value, bool):
                        # Direct boolean value - find matching choice
                        choice_text = 'True' if answer_value else 'False'
                        selected_choice = question.choices.filter(text__iexact=choice_text).first()
                        answer.boolean_answer = answer_value
                    elif isinstance(answer_value, str):
                        if answer_value.lower() in ['true', 'false']:
                            # String value ('true' or 'false')
                            answer.boolean_answer = answer_value.lower() == 'true'
                            selected_choice = question.choices.filter(text__iexact=answer_value).first()
                        else:
                            # Might be a choice ID as string
                            try:
                                choice_id = int(answer_value)
                                selected_choice = Choice.objects.get(id=choice_id, question=question)
                                answer.boolean_answer = selected_choice.text.lower() == 'true'
                            except (ValueError, TypeError, Choice.DoesNotExist):
                                answer.boolean_answer = answer_value.lower() == 'true'
                    elif isinstance(answer_value, int):
                        # Could be choice ID or 0/1 boolean
                        try:
                            # First try as choice ID
                            selected_choice = Choice.objects.get(id=answer_value, question=question)
                            answer.boolean_answer = selected_choice.text.lower() == 'true'
                            logger.debug(f"True/False: Found choice {selected_choice.id} - '{selected_choice.text}'")
                        except Choice.DoesNotExist:
                            # If not a valid choice ID, treat as 1=true, 0=false
                            answer.boolean_answer = answer_value == 1
                            choice_text = 'True' if answer.boolean_answer else 'False'
                            selected_choice = question.choices.filter(text__iexact=choice_text).first()
                    else:
                        # Any other value, convert using Python's bool()
                        answer.boolean_answer = bool(answer_value)
                        choice_text = 'True' if answer.boolean_answer else 'False'
                        selected_choice = question.choices.filter(text__iexact=choice_text).first()
                    
                    # Add the selected choice to the answer
                    if selected_choice:
                        answer.selected_choices.add(selected_choice)
                        logger.debug(f"True/False: Selected choice {selected_choice.id} - '{selected_choice.text}' (is_correct={selected_choice.is_correct})")
                        
                        # For True/False, simply check if the selected choice is marked as correct
                        if selected_choice.is_correct:
                            answer.points_earned = question.points
                            answer.is_correct = True
                            logger.info(f"True/False question {question.id} marked as CORRECT - selected correct choice")
                        else:
                            answer.is_correct = False
                            answer.points_earned = 0
                            logger.info(f"True/False question {question.id} marked as INCORRECT - selected wrong choice")
                    else:
                        # Fallback: Check using correct_answer field if no choice found
                        logger.warning(f"No choice found for true/false answer, using correct_answer field")
                        correct_answer = question.correct_answer
                        if correct_answer is not None:
                            # Convert correct_answer to boolean if it's a string
                            if isinstance(correct_answer, str):
                                correct_answer_bool = correct_answer.lower() in ['true', '1', 'yes']
                            elif isinstance(correct_answer, bool):
                                correct_answer_bool = correct_answer
                            else:
                                try:
                                    correct_answer_bool = bool(int(correct_answer))
                                except (ValueError, TypeError):
                                    correct_answer_bool = bool(correct_answer)
                            
                            logger.debug(f"Comparing answer {answer.boolean_answer} with correct {correct_answer_bool}")
                            
                            if answer.boolean_answer == correct_answer_bool:
                                answer.points_earned = question.points
                                answer.is_correct = True
                                logger.info(f"True/False question {question.id} marked as CORRECT")
                            else:
                                answer.is_correct = False
                                answer.points_earned = 0
                                logger.info(f"True/False question {question.id} marked as INCORRECT")
                        else:
                            # No way to determine correctness
                            logger.error(f"Cannot determine correct answer for true/false question {question.id}")
                            answer.is_correct = False
                            answer.points_earned = 0
                            
                except Exception as e:
                    logger.error(f"Error processing true/false answer: {str(e)}", exc_info=True)
                    # Don't award points if there was an error processing the answer
                    answer.is_correct = False
                    answer.points_earned = 0
        
        answer.save()
        total_points += question.points
        earned_points += answer.points_earned
        logger.debug(f"Question {question.id}: worth {question.points} points, earned {answer.points_earned}. Running totals: {earned_points}/{total_points}")
    
    try:
        # Handle edge case where no questions were answered
        question_count = quiz.questions.count()
        if question_count == 0:
            logger.warning("Quiz has no questions")
            attempt.score = 0
            attempt.total_points = 0
            attempt.total_questions = 0
            attempt.percentage = 0
            attempt.status = 'submitted'
            attempt.completed_at = timezone.now()
            attempt.save()
            
            logger.warning(f"Quiz {quiz_id} has no questions. Setting score to 0/0.")
            
            return JsonResponse({
                'success': True,
                'score': 0,
                'total': 0,
                'percentage': 0,
                'passed': False,
                'message': 'Quiz has no questions',
                'redirect': reverse('academic_integration:quiz_result', args=[quiz.id])
            })
            
        # Store the total questions count
        attempt.total_questions = question_count
        
        # Handle edge case where no points were earned or total_points is 0
        if total_points == 0:
            logger.warning(f"Quiz {quiz.id} has total_points=0. Setting percentage to 0.")
            attempt.percentage = 0
        elif earned_points == 0:
            logger.warning(f"Student earned 0 points on quiz {quiz.id} out of {total_points} possible points.")
            attempt.percentage = 0
        else:
            attempt.percentage = (earned_points / total_points * 100)
        
        # Update attempt with score
        attempt.score = earned_points
        attempt.total_points = total_points
        attempt.percentage = (earned_points / total_points * 100) if total_points > 0 else 0
        
        # Check if no answers were recorded
        if attempt.answers.count() == 0:
            logger.warning(f"No answers were recorded for quiz attempt {attempt.id}. This might indicate a submission issue.")
            messages.warning(request, "Note: No answers were recorded for this quiz attempt. Your score may be affected.")
        
        # Check if passed
        if attempt.percentage >= quiz.passing_score:
            attempt.passed = True
        
        # Check if quiz has any text questions that require manual grading
        has_text_questions = quiz.questions.filter(question_type='text').exists()
        
        # If no text questions, automatically mark as graded since all questions are auto-graded
        if not has_text_questions:
            attempt.status = 'graded'
            logger.info(f"Quiz {quiz.id} has no text questions. Automatically marking attempt {attempt.id} as graded.")
        
        attempt.save()
        
        logger.debug(f"Quiz submission successful - Score: {attempt.score}/{attempt.total_points} ({attempt.percentage}%), Status: {attempt.status}")
        
        # Store quiz results as tutorial marks if applicable
        if quiz.quiz_type == 'tutorial' and quiz.tutorial_number and quiz.course_id:
            try:
                # Format the score for academic analyzer - ensure it's a correct scaled value
                tutorial_number = quiz.tutorial_number
                
                # Calculate the scaled score properly
                # If total points is 0, use 0 as the scaled score
                # Otherwise, calculate what percentage of total points were earned and scale to 0-10
                if total_points > 0:
                    # Improved calculation: directly use earned_points/total_points ratio 
                    # multiplied by 10 to get a score out of 10
                    scaled_score = (earned_points / total_points) * 10
                    logger.info(f"Calculated tutorial mark: {earned_points}/{total_points} * 10 = {scaled_score}")
                else:
                    scaled_score = 0
                    logger.warning("Total points is 0, setting scaled score to 0")
                
                # Get the teacher's email associated with the course or from session
                teacher_email = None
                
                # First check if we have a teacher linked to the quiz directly
                if quiz.created_by and quiz.created_by.email:
                    teacher_email = quiz.created_by.email
                    logger.info(f"Using quiz creator's email for update: {teacher_email}")
                
                # If no teacher email yet, try to get course instructor from API
                if not teacher_email:
                    try:
                        # Try to get course details to find the instructor
                        course_response = requests.get(
                            f"{api_base_url()}/staff/course-detail",
                            params={"courseId": quiz.course_id},
                            timeout=5,
                        )
                        
                        if course_response.ok:
                            course_data = _safe_json(course_response)
                            if course_data.get("success"):
                                # Use instructor email if available
                                teacher_email = course_data.get("instructorEmail")
                                if teacher_email:
                                    logger.info(f"Found instructor email for course {quiz.course_id}: {teacher_email}")
                    except Exception as e:
                        logger.warning(f"Failed to get instructor email from API: {str(e)}")
                
                # If still no teacher email, check if there's a staff email in the session
                if not teacher_email and request.session.get('staff_email'):
                    teacher_email = request.session.get('staff_email')
                    logger.info(f"Using staff email from session: {teacher_email}")
                
                # As a last resort, use the course ID to generate a default teacher email
                if not teacher_email:
                    # Default format based on course ID - e.g. "teacher_COURSE101@psgtech.ac.in"
                    teacher_email = f"teacher_{quiz.course_id.lower()}@psgtech.ac.in"
                    logger.warning(f"No teacher email found, using generated fallback: {teacher_email}")
                
                # Call Academic Analyzer API to update tutorial marks using the staff/update-student-marks endpoint
                update_marks_response = requests.post(
                    f"{api_base_url()}/staff/update-student-marks",
                    json={
                        'studentId': student_roll_number,
                        'courseId': quiz.course_id,
                        'teacherEmail': teacher_email,
                        'marks': {
                            f'tutorial{tutorial_number}': scaled_score
                        }
                    },
                    timeout=10  # Increased timeout for better reliability
                )
                
                if update_marks_response.ok:
                    marks_data = _safe_json(update_marks_response)
                    if marks_data.get('success'):
                        logger.info(f"Successfully updated tutorial marks for student {student_roll_number} in course {quiz.course_id}, tutorial {tutorial_number}: {scaled_score}")
                        # Mark the attempt as synced with Academic Analyzer
                        attempt.marks_synced = True
                        attempt.last_sync_at = timezone.now()
                        attempt.save()
                        logger.info(f"Marked quiz attempt {attempt.id} as synced with Academic Analyzer")
                    else:
                        error_msg = marks_data.get('message', 'Unknown error')
                        logger.warning(f"Failed to update tutorial marks: {error_msg}")
                        # Add diagnostic info to the success response
                        messages.warning(request, f"Note: Your quiz was submitted successfully, but there was an error syncing the marks to Academic Analyzer: {error_msg}. Please inform your instructor.")
                else:
                    logger.warning(f"Failed to update tutorial marks. API responded with status code: {update_marks_response.status_code}")
                    # Add diagnostic info to the success response
                    messages.warning(request, f"Note: Your quiz was submitted successfully, but there was an error syncing the marks to Academic Analyzer (Status: {update_marks_response.status_code}). Please inform your instructor.")
            except requests.RequestException as e:
                logger.exception(f"Failed to update tutorial marks: {e}")
                # Add diagnostic info to the success response
                messages.warning(request, "Note: Your quiz was submitted successfully, but there was a connection error syncing the marks to Academic Analyzer. Please inform your instructor.")
                
        # Return success with redirect to results
        return JsonResponse({
            'success': True, 
            'score': attempt.score,
            'total': attempt.total_points,
            'percentage': attempt.percentage,
            'passed': attempt.passed,
            'tutorial_sync_status': attempt.marks_synced,
            'redirect': reverse('academic_integration:quiz_result', args=[quiz.id])
        })
    except Exception as e:
        logger.error(f"Error saving quiz attempt: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f"Error saving quiz results: {str(e)}"
        }, status=500)
def staff_logout(request: HttpRequest) -> HttpResponse:
	request.session.flush()
	messages.success(request, "You have been logged out.")
	return redirect("academic_integration:staff_login")


def quiz_availability_info(request: HttpRequest, quiz_id: int) -> HttpResponse:
    """
    View for students to get detailed information about quiz availability.
    This is helpful for diagnosing "Quiz not yet available" issues.
    """
    from quiz.models import Quiz
    from django.shortcuts import get_object_or_404
    from django.conf import settings
    
    # Ensure student is logged in
    student_roll_number = request.session.get('student_roll_number')
    
    if not student_roll_number:
        messages.error(request, "Please log in to continue.")
        return redirect("academic_integration:student_login")
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    # Check availability using the model method
    is_available = quiz.is_available
    is_visible, availability_reason = quiz.debug_visibility_status()
    
    # Check if dates are timezone-naive
    start_date_naive = quiz.start_date and timezone.is_naive(quiz.start_date)
    complete_by_naive = quiz.complete_by_date and timezone.is_naive(quiz.complete_by_date)
    
    context = {
        'quiz': quiz,
        'is_available': is_available,
        'availability_reason': availability_reason,
        'start_date_naive': start_date_naive,
        'complete_by_naive': complete_by_naive,
        'current_time': timezone.now(),
        'time_zone': settings.TIME_ZONE,
        'student_roll_number': student_roll_number,
        'student_name': request.session.get('student_name', student_roll_number),
    }
    
    return render(request, "academic_integration/quiz_availability_info.html", context)


def student_quiz_dashboard(request: HttpRequest) -> HttpResponse:
    """
    View for students to see available quizzes and their past attempts.
    """
    from quiz.models import Quiz, QuizAttempt
    from django.db.models import Q, Count, Max
    import logging
    
    # Set up logging
    logger = logging.getLogger(__name__)
    
    # Ensure student is logged in
    student_roll_number = request.session.get("student_roll_number")
    if not student_roll_number:
        messages.info(request, "Please log in to continue.")
        return redirect("academic_integration:student_login")
    
    # Debug - log student info
    logger.debug(f"Loading quizzes for student: {student_roll_number}")
    
    # Check if there's an unavailable quiz ID in the session
    unavailable_quiz_id = request.session.pop('unavailable_quiz_id', None)
    if unavailable_quiz_id:
        # Add an additional message with the link to the availability info
        messages.info(request, f"For more information about why this quiz is not available, <a href='{reverse('academic_integration:quiz_availability_info', args=[unavailable_quiz_id])}'>click here</a>")
    
    # Get student's enrolled courses
    enrolled_courses = []
    courses_data = []
    api_error = None
    
    try:
        response = requests.get(
            f"{api_base_url()}/student/dashboard",
            params={"rollno": student_roll_number},
            timeout=5,
        )
        if response.ok:
            data = _safe_json(response)
            if data.get('success'):
                courses_data = data.get('courses', [])
                enrolled_courses = [course['courseId'] for course in courses_data]
                logger.debug(f"Retrieved {len(enrolled_courses)} courses: {enrolled_courses}")
            else:
                api_error = "Failed to fetch course data from academic API."
                logger.warning(f"API error: {data.get('message', 'Unknown error')}")
        else:
            api_error = f"API responded with status code: {response.status_code}"
            logger.warning(f"API error: {response.status_code} - {response.text}")
    except requests.RequestException as e:
        api_error = "Failed to connect to academic API."
        logger.exception(f"Failed to fetch courses for student quiz dashboard: {e}")
    
    # Create a course lookup dictionary for faster access
    course_lookup = {course['courseId']: course for course in courses_data}
    
    # Check if filtering by course
    course_filter = request.GET.get('course_id')
    if course_filter:
        logger.debug(f"Filtering by course ID: {course_filter}")
    
    # Get today's date for filtering active quizzes
    today = timezone.now()
    
    # Try to get all quizzes regardless of API status to ensure we display something
    available_quizzes = []
    
    # Try direct database query first to show only quizzes for enrolled courses
    try:
        # Check if there are any quizzes in the database, if not, create a demo quiz
        if Quiz.objects.count() == 0:
            create_demo_quiz()
        
        # Apply filters to show only quizzes for enrolled courses
        quiz_filter = Q()
        
        # Filter to only show quizzes for courses the student is enrolled in
        if enrolled_courses:
            quiz_filter &= Q(course_id__in=enrolled_courses)
        
        # If filtering by specific course, add that filter
        if course_filter:
            quiz_filter &= Q(course_id=course_filter)
        
        # Get quizzes from the database for enrolled courses
        available_quizzes = Quiz.objects.filter(quiz_filter).prefetch_related('questions').order_by('-created_at')
        logger.debug(f"Direct DB query found {len(available_quizzes)} quizzes for enrolled courses")
        
        # Log all quizzes for debugging
        for q in available_quizzes[:10]:  # Limit to first 10 to avoid flooding logs
            logger.debug(f"Quiz found: ID={q.id}, Title={q.title}, Active={q.is_active}, Course={q.course_id}")
    except Exception as e:
        logger.exception(f"Error querying quizzes directly: {e}")
    
    # Process all quizzes for display
    processed_quizzes = []
	
    for quiz in available_quizzes:
        try:
            # Count total questions
            quiz.question_count = quiz.questions.count()
            
            # Check if student has attempted this quiz
            attempt = QuizAttempt.objects.filter(
                quiz=quiz,
                user__username=student_roll_number
            ).order_by('-started_at').first()
            
            quiz.attempt = attempt
            
            # Add course name directly to quiz object
            if quiz.course_id in course_lookup:
                quiz.course_name = course_lookup[quiz.course_id]['courseName']
                quiz.course_code = course_lookup[quiz.course_id]['courseCode']
            else:
                quiz.course_name = f"Course {quiz.course_id}" if quiz.course_id else "General Quiz"
                quiz.course_code = quiz.course_id or ""
            
            # Check if quiz can be attempted (not completed or allowed for retake)
            quiz.can_attempt = (not attempt or not attempt.completed_at or quiz.allow_retake)
            quiz.duration_minutes = quiz.duration_minutes or 30
			
            # Only add if it has questions
            if quiz.question_count > 0:
                processed_quizzes.append(quiz)
            
        except Exception as e:
            logger.exception(f"Error processing quiz {quiz.id}: {e}")
    
    # Get completed quiz attempts
    completed_attempts = QuizAttempt.objects.filter(
        user__username=student_roll_number,
        completed_at__isnull=False
    ).select_related('quiz').order_by('-completed_at')
    
    logger.debug(f"Final processed quizzes: {len(processed_quizzes)}")
    
    context = {
        'available_quizzes': processed_quizzes,
        'completed_attempts': completed_attempts,
        'student_roll_number': student_roll_number,
        'student_name': request.session.get('student_name', student_roll_number),
        'courses': courses_data,
        'selected_course_id': course_filter,
        'api_error': api_error,
        'quiz_count': len(processed_quizzes),
        'enrolled_courses': enrolled_courses,  # This is now properly passed to the template
        'debug_mode': True,  # Enable debug mode in template
    }
    
    return render(request, "academic_integration/student_quiz_dashboard.html", context)


def quiz_detail(request: HttpRequest, quiz_id: int) -> HttpResponse:
    """
    View for students to take a quiz.
    """
    from quiz.models import Quiz, QuizAttempt, User
    from academic_integration.models import Student
    from django.shortcuts import get_object_or_404
    
    # Ensure student is logged in
    student_roll_number = request.session.get("student_roll_number")
    student_id = request.session.get("student_id")
    
    if not student_roll_number:
        messages.info(request, "Please log in to continue.")
        return redirect("academic_integration:student_login")
    
    # Get the quiz
    quiz = get_object_or_404(Quiz, pk=quiz_id, is_active=True)
	
    # Check if student is enrolled in the course for this quiz
    if quiz.course_id:
        enrolled_courses = []
        try:
            response = requests.get(
                f"{api_base_url()}/student/dashboard",
                params={"rollno": student_roll_number},
                timeout=5,
            )
            if response.ok:
                data = _safe_json(response)
                if data.get('success'):
                    enrolled_courses = [course['courseId'] for course in data.get('courses', [])]
        except requests.RequestException:
            logger.exception("Failed to fetch courses for quiz detail")
        
        # Enforce enrollment check to prevent access to quizzes from courses the student is not enrolled in
        if quiz.course_id not in enrolled_courses:
            logger.warning(f"Student {student_roll_number} attempted to access quiz {quiz_id} for course {quiz.course_id} but is not enrolled. Enrolled courses: {enrolled_courses}")
            messages.error(request, "You cannot access this quiz because you are not enrolled in the course.")
            return redirect("academic_integration:student_quiz_dashboard")
        logger.info(f"Student {student_roll_number} accessing quiz {quiz_id} for course {quiz.course_id}. Enrolled courses: {enrolled_courses}")
    
    # Check if quiz is available using the quiz model's is_available property
    if not quiz.is_available:
        is_visible, reason = quiz.debug_visibility_status()
        messages.error(request, f"{reason}")
        # Store quiz ID in session for availability info link
        request.session['unavailable_quiz_id'] = quiz_id
        return redirect("academic_integration:student_quiz_dashboard")
    
    # Get or create the student user
    student_user, created = User.objects.get_or_create(
        username=student_roll_number,
        defaults={
            'email': f"{student_roll_number}@psgtech.ac.in",
            'role': 'student'
        }
    )
    
    # Get or create student profile
    student, created = Student.objects.get_or_create(
        user=student_user,
        defaults={
            'student_id': student_id or student_roll_number  # Use academic analyzer ID if available
        }
    )
    
    # Check for existing attempts
    attempt = QuizAttempt.objects.filter(
        quiz=quiz,
        user=student.user
    ).order_by('-started_at').first()
    
    # If completed and no retakes allowed, show results
    if attempt and attempt.completed_at and not quiz.allow_retake:
        return redirect("academic_integration:quiz_result", quiz_id=quiz_id)
    
    context = {
        'quiz': quiz,
        'student_roll_number': student_roll_number,
        'student_name': request.session.get('student_name', student_roll_number),
    }
    
    return render(request, "academic_integration/quiz_detail.html", context)


def quiz_result(request: HttpRequest, quiz_id: int) -> HttpResponse:
    """
    View for students to see their quiz results.
    """
    from quiz.models import Quiz, QuizAttempt, User
    from academic_integration.models import Student
    from django.shortcuts import get_object_or_404
	
    # Ensure student is logged in
    student_roll_number = request.session.get("student_roll_number")
    
    if not student_roll_number:
        messages.info(request, "Please log in to continue.")
        return redirect("academic_integration:student_login")
    
    # Get the quiz
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    # Check if student is enrolled in the course for this quiz
    if quiz.course_id:
        enrolled_courses = []
        try:
            response = requests.get(
                f"{api_base_url()}/student/dashboard",
                params={"rollno": student_roll_number},
                timeout=5,
            )
            if response.ok:
                data = _safe_json(response)
                if data.get('success'):
                    enrolled_courses = [course['courseId'] for course in data.get('courses', [])]
        except requests.RequestException:
            logger.exception("Failed to fetch courses for quiz result")
        
        # Enforce enrollment check to prevent access to quizzes from courses the student is not enrolled in
        if quiz.course_id not in enrolled_courses:
            logger.warning(f"Student {student_roll_number} attempted to access quiz result {quiz_id} for course {quiz.course_id} but is not enrolled")
            messages.error(request, "You cannot access this quiz result because you are not enrolled in the course.")
            return redirect("academic_integration:student_quiz_dashboard")
    
    # Get the student
    student = get_object_or_404(Student, user__username=student_roll_number)
    
    # Get the quiz attempt
    quiz_attempt = QuizAttempt.objects.filter(
        quiz=quiz,
        user=student.user,
        completed_at__isnull=False
    ).order_by('-completed_at').first()
    
    # If no completed attempt, redirect to the quiz
    if not quiz_attempt:
        return redirect("academic_integration:quiz_detail", quiz_id=quiz_id)
    
    # If results are not shown and student is not allowed to review, redirect to dashboard
    if not quiz.show_results and not quiz.allow_review:
        messages.info(request, "Results for this quiz are not available for review.")
        return redirect("academic_integration:student_quiz_dashboard")
    
    # Calculate percentage score (already stored in quiz_attempt.percentage)
    percentage = quiz_attempt.percentage
    
    # Get the total number of questions
    total_questions = quiz.questions.count()
    
    # Make sure quiz_attempt.total_questions is also updated if not set
    if not quiz_attempt.total_questions or quiz_attempt.total_questions == 0:
        quiz_attempt.total_questions = total_questions
        quiz_attempt.save()
        logger.info(f"Updated total_questions for quiz_attempt {quiz_attempt.id} to {total_questions}")
    
    logger.info(f"Quiz {quiz_id} result page - Score: {quiz_attempt.score}/{quiz_attempt.total_points}, Percentage: {percentage}%, Questions: {total_questions}")
    
    # If the quiz attempt has no answers but the quiz has questions, add a warning
    has_no_answers = quiz_attempt.answers.count() == 0 and total_questions > 0
    
    # Create a dictionary mapping question IDs to answers for easier template access
    question_answers = {}
    for answer in quiz_attempt.answers.all():
        question_answers[answer.question.id] = answer
    
    context = {
        'quiz': quiz,
        'quiz_attempt': quiz_attempt,
        'percentage': percentage,
        'total_questions': total_questions,
        'has_no_answers': has_no_answers,
        'question_answers': question_answers,
        'student_roll_number': student_roll_number,
        'student_name': request.session.get('student_name', student_roll_number),
    }
    
    return render(request, "academic_integration/quiz_result.html", context)


def home(request: HttpRequest) -> HttpResponse:
    """
    Home page view that serves as the main entry point for the application.
    Redirects to the appropriate dashboard if the user is already logged in.
    """
    # If staff is logged in, redirect to staff dashboard
    if request.session.get("staff_email"):
        return redirect("academic_integration:staff_dashboard")
    
    # If student is logged in, redirect to student dashboard
    if request.session.get("student_roll_number"):
        return redirect("academic_integration:student_dashboard")
    
    # Otherwise show the landing page
    return render(request, "academic_integration/home.html")


def student_login(request: HttpRequest) -> HttpResponse:
	"""
	Student login view that authenticates students using the Academic Analyzer API.
	"""
	if request.session.get("student_roll_number"):
		return redirect("academic_integration:student_dashboard")

	form = StudentLoginForm(request.POST or None)

	if request.method == "POST" and form.is_valid():
		payload = form.cleaned_data
		try:
			response = requests.post(
				f"{api_base_url()}/student/auth",
				json={"rollno": payload["rollno"], "password": payload["password"]},
				timeout=5,
			)
		except requests.RequestException:
			logger.exception("Student auth request failed")
			form.add_error(None, "Cannot reach Academic Analyzer API right now. Please try again later.")
		else:
			body = _safe_json(response)
			if response.ok and body.get("success"):
				request.session["student_roll_number"] = body.get("rollno", payload["rollno"])
				request.session["student_id"] = body.get("studentId")
				request.session["student_name"] = body.get("name") or body.get("rollno") or payload["rollno"]
				messages.success(request, "Logged in successfully.")
				return redirect("academic_integration:student_dashboard")
			error_message = body.get("message", "Invalid credentials. Please try again.")
			form.add_error(None, error_message)

	return render(request, "academic_integration/student_login.html", {"form": form})


def student_dashboard(request: HttpRequest) -> HttpResponse:
	"""
	Student dashboard view that displays enrolled courses, academic performance, and active quizzes.
	"""
	from quiz.models import Quiz, QuizAttempt
	from django.db.models import Q, Count, Max
	
	# Add debugging session info
	logger.info(f"Student dashboard accessed. Session data: {request.session.items()}")
	
	student_roll_number = request.session.get("student_roll_number")
	if not student_roll_number:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:student_login")

	courses: list[Dict[str, Any]] = []
	performance: list[Dict[str, Any]] = []
	api_error: str | None = None
	enrolled_courses = []

	try:
		response = requests.get(
			f"{api_base_url()}/student/dashboard",
			params={"rollno": student_roll_number},
			timeout=5,
		)
	except requests.RequestException:
		logger.exception("Failed to load student dashboard data")
		api_error = "Could not reach Academic Analyzer. Please refresh the page later."
	else:
		body = _safe_json(response)
		if response.ok and body.get("success"):
			courses = body.get("courses", [])
			enrolled_courses = [course['courseId'] for course in courses]
			performance = body.get("performance", [])
			if body.get("name"):
				request.session["student_name"] = body["name"]
		else:
			api_error = body.get("message", "Unable to load dashboard data.")

	# Get today's date for filtering active quizzes
	today = timezone.now()
	
	# Create a course lookup dictionary for faster access
	course_lookup = {course['courseId']: course for course in courses}
	
	# Get available quizzes using both the direct database query and our new API
	available_quizzes = []
	
	# First, try to get quizzes using our new API endpoint
	try:
		from rest_framework.test import APIRequestFactory
		factory = APIRequestFactory()
		from quiz.views import get_active_quizzes
		api_request = factory.get(f'/quiz/active-quizzes/student/{student_roll_number}/')
		api_request.session = request.session
		response = get_active_quizzes(api_request, student_roll_number)
		
		if response.status_code == 200:
			api_quizzes = response.data.get('quizzes', [])
			
			# If we have quizzes from the API, use those
			if api_quizzes:
				# Get the full quiz objects from the database
				quiz_ids = [q['id'] for q in api_quizzes]
				available_quizzes = Quiz.objects.filter(id__in=quiz_ids).prefetch_related('questions').order_by('-created_at')
				
				# Enrich with attempt information
				for quiz in available_quizzes:
					# Count total questions
					quiz.question_count = quiz.questions.count()
					
					# Check if student has attempted this quiz
					attempt = QuizAttempt.objects.filter(
						quiz=quiz,
						user__username=student_roll_number
					).order_by('-started_at').first()
					
					quiz.attempt = attempt
					
					# Add course name directly to quiz object
					if quiz.course_id in course_lookup:
						quiz.course_name = course_lookup[quiz.course_id]['courseName']
						quiz.course_code = course_lookup[quiz.course_id].get('courseCode', '')
	except Exception as e:
		logger.exception(f"Error using active quizzes API: {e}")
		
	# UPDATED: Use the direct database query approach with proper filtering
	# Query only active quizzes for enrolled courses to begin with
	query_filter = Q(is_active=True)
	
	# Add filter for enrolled courses directly in the database query
	if enrolled_courses:
		query_filter &= Q(course_id__in=enrolled_courses)
		# Get active quizzes for enrolled courses
		available_quizzes = Quiz.objects.filter(query_filter).prefetch_related('questions').order_by('-created_at')
		logger.info(f"Found {available_quizzes.count()} active quizzes for enrolled courses: {enrolled_courses}")
	else:
		# If no enrolled courses, return an empty queryset
		available_quizzes = Quiz.objects.none()
		logger.info("Student has no enrolled courses, returning empty quiz set")
	
	# Log the number of quizzes found
	logger.info(f"Found {available_quizzes.count()} active quizzes for enrolled courses")
	
	# Create a list to store processed quizzes
	processed_quizzes = []
	
	# Process all quizzes - whether from API or direct query
	for quiz in available_quizzes:
		try:
			# IMPROVED: Log each quiz being processed for debugging
			logger.info(f"Processing quiz ID: {quiz.id}, Title: '{quiz.title}', Course ID: '{quiz.course_id}', Active: {quiz.is_active}")
			
			# Use the new debug method to check visibility status
			is_visible, visibility_reason = quiz.debug_visibility_status()
			quiz.visibility_status = visibility_reason
			
			# Skip quizzes with no questions (handled in debug_visibility_status)
			if not is_visible and "no questions" in visibility_reason:
				logger.info(f"Skipping quiz ID {quiz.id} - has no questions")
				continue
			
			# For debugging, we'll process ALL quizzes regardless of visibility status
			quiz.question_count = quiz.questions.count()
			
			# Check if student has attempted this quiz
			attempt = QuizAttempt.objects.filter(
				quiz=quiz,
				user__username=student_roll_number
			).order_by('-started_at').first()
			
			quiz.attempt = attempt
			
			# Add course name directly to quiz object
			if quiz.course_id in course_lookup:
				quiz.course_name = course_lookup[quiz.course_id]['courseName']
				quiz.course_code = course_lookup[quiz.course_id].get('courseCode', '')
				logger.info(f"Quiz {quiz.id} matched to enrolled course '{quiz.course_name}' ({quiz.course_id})")
			else:
				quiz.course_name = f"Course {quiz.course_id}" if quiz.course_id else "General Quiz"
				quiz.course_code = quiz.course_id or ""
				logger.info(f"Quiz {quiz.id} has course_id '{quiz.course_id}' which is not in enrolled courses")
			
			# Check if quiz can be attempted (not completed or allowed for retake)
			quiz.can_attempt = (not attempt or not attempt.completed_at or quiz.allow_retake)
			
			# Add debugging info to the quiz object
			quiz.debug = {
				"is_visible": is_visible,
				"visibility_reason": visibility_reason,
				"in_enrolled_course": quiz.course_id in enrolled_courses if quiz.course_id else "No course ID",
				"has_questions": quiz.question_count > 0,
				"is_available": quiz.is_available,
				"today": timezone.now(),
			}
			
			# Add the processed quiz to our list - in debug mode, show all quizzes
			processed_quizzes.append(quiz)
			logger.info(f"Quiz {quiz.id} added to processed_quizzes list. Visible: {is_visible}, Reason: {visibility_reason}")
		except Exception as e:
			logger.exception(f"Error processing quiz {quiz.id}: {e}")
	
	# Get recent completed quiz attempts
	recent_attempts = QuizAttempt.objects.filter(
		user__username=student_roll_number,
		completed_at__isnull=False
	).select_related('quiz').order_by('-completed_at')[:3]  # Limit to 3 most recent completed attempts
	
	# Add course name to each attempt
	for attempt in recent_attempts:
		if attempt.quiz.course_id in course_lookup:
			attempt.course_name = course_lookup[attempt.quiz.course_id]['courseName']
		else:
			attempt.course_name = f"Course {attempt.quiz.course_id}" if attempt.quiz.course_id else "General Quiz"
	
	logger.info(f"Student dashboard for {student_roll_number}: Found {len(processed_quizzes)} quizzes to display")
	
	context = {
		"student_name": request.session.get("student_name") or student_roll_number,
		"student_roll_number": student_roll_number,
		"courses": courses,
		"performance": performance,
		"api_error": api_error,
		"available_quizzes": processed_quizzes,  # Use processed_quizzes instead of available_quizzes
		"recent_attempts": recent_attempts,
		"quiz_count": len(processed_quizzes),  # Add a count for easier debugging
		"enrolled_courses": enrolled_courses,  # Pass enrolled courses for debugging
		"debug_mode": True,  # Enable debug mode in template
	}
	return render(request, "academic_integration/student_dashboard.html", context)


def course_detail(request: HttpRequest, course_id: str) -> HttpResponse:
	"""
	Detailed view of a single course for students, showing progress, quizzes, and performance history.
	"""
	from quiz.models import Quiz, QuizAttempt
	from django.db.models import Q, Avg, Count
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	
	student_roll_number = request.session.get("student_roll_number")
	if not student_roll_number:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:student_login")
	
	# Get course details from Academic Analyzer API
	course = {}
	performance = []
	api_error = None
	
	# Additional variables for marks data
	student_marks = {}
	component_details = []
	
	# Verify student is enrolled in this course
	enrolled_courses = []
	try:
		response = requests.get(
			f"{api_base_url()}/student/dashboard",
			params={"rollno": student_roll_number},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get('success'):
				courses_data = data.get('courses', [])
				enrolled_courses = [course['courseId'] for course in courses_data]
				
				# Find the specific course in the list
				for c in courses_data:
					if c['courseId'] == course_id:
						course = c
						break
				
				performance = data.get('performance', [])
				# Filter performance for this course only
				performance = [p for p in performance if p.get('courseId') == course_id]
	except requests.RequestException:
		logger.exception("Failed to fetch course details for student")
		api_error = "Could not reach Academic Analyzer API. Please try again later."
	
	# Check if student is enrolled in this course
	if course_id not in enrolled_courses:
		messages.error(request, "You are not enrolled in this course.")
		return redirect("academic_integration:student_dashboard")
		
	# Get detailed marks from Academic Analyzer API
	try:
		marks_response = requests.get(
			f"{api_base_url()}/student/course-marks",
			params={"rollno": student_roll_number, "courseId": course_id},
			timeout=5,
		)
		if marks_response.ok:
			marks_data = _safe_json(marks_response)
			if marks_data.get('success'):
				# Update course info if available
				if marks_data.get('courseName'):
					course['courseName'] = marks_data.get('courseName')
				if marks_data.get('courseCode'):
					course['courseCode'] = marks_data.get('courseCode')
				
				student_marks = marks_data.get('marks', {})
				component_details = marks_data.get('components', [])
				
				logger.info(f"Retrieved {len(component_details)} component details for student {student_roll_number} in course {course_id}")
			else:
				logger.warning(f"API returned success=false for marks: {marks_data.get('message')}")
				# Not showing error to user since this is an additional feature
	except requests.RequestException as e:
		logger.exception(f"Error fetching course marks data: {e}")
		# Not showing error to user since this is an additional feature
	
	# Get today's date for filtering active quizzes
	today = timezone.now()
	
	# Get all quizzes for this course
	quizzes = Quiz.objects.filter(
		course_id=course_id,
		is_active=True
	).prefetch_related('questions').order_by('-created_at')
	
	# Get attempts by this student for the quizzes
	for quiz in quizzes:
		# Count total questions
		quiz.question_count = quiz.questions.count()
		
		# Check if student has attempted this quiz
		attempt = QuizAttempt.objects.filter(
			quiz=quiz,
			user__username=student_roll_number
		).order_by('-started_at').first()
		
		quiz.attempt = attempt
	
	# Filter active quizzes (not expired)
	active_quizzes = [q for q in quizzes if (not q.complete_by_date or q.complete_by_date >= today)]
	
	# Get completed quiz attempts for this course
	completed_attempts = QuizAttempt.objects.filter(
		user__username=student_roll_number,
		completed_at__isnull=False,
		quiz__course_id=course_id
	).select_related('quiz').order_by('-completed_at')
	
	# Calculate course progress metrics
	total_quizzes = len(quizzes)
	completed_quizzes = sum(1 for q in quizzes if any(a.quiz.id == q.id and a.completed_at for a in completed_attempts))
	
	completion_percentage = (completed_quizzes / total_quizzes * 100) if total_quizzes > 0 else 0
	quiz_completion_percentage = (completed_quizzes / total_quizzes * 100) if total_quizzes > 0 else 0
	
	# Calculate average score
	average_score = 0
	if completed_attempts:
		average_score = completed_attempts.aggregate(Avg('percentage'))['percentage__avg'] or 0
	
	# Determine course grade based on average score
	course_grade = 'N/A'
	if average_score >= 90:
		course_grade = 'A'
	elif average_score >= 80:
		course_grade = 'B'
	elif average_score >= 70:
		course_grade = 'C'
	elif average_score >= 60:
		course_grade = 'D'
	elif average_score > 0:
		course_grade = 'F'
	
	context = {
		"student_name": request.session.get("student_name") or student_roll_number,
		"student_roll_number": student_roll_number,
		"course": course,
		"performance": performance,
		"api_error": api_error,
		"quizzes": quizzes,
		"active_quizzes": active_quizzes,
		"completed_attempts": completed_attempts,
		"completion_percentage": round(completion_percentage),
		"quiz_completion_percentage": round(quiz_completion_percentage),
		"average_score": average_score,
		"course_grade": course_grade,
		"total_quizzes": total_quizzes,
		"completed_quizzes": completed_quizzes,
		"course_performance": bool(completed_attempts),  # Whether to show performance metrics
		# Add marks data
		"student_marks": student_marks,
		"component_details": component_details,
		"has_marks_data": bool(component_details),
	}
	
	return render(request, "academic_integration/course_detail.html", context)
	
	
def student_active_quizzes(request: HttpRequest) -> HttpResponse:
	"""
	Returns active quizzes for the logged-in student.
	Used as an API endpoint by the student dashboard.
	"""
	from quiz.models import Quiz, QuizAttempt
	from django.db.models import Q
	from django.http import JsonResponse
	from django.utils import timezone
	
	# Check if student is logged in
	if not request.session.get("student_roll_number"):
		return JsonResponse({"error": "Not logged in"}, status=401)
	
	student_roll_number = request.session.get("student_roll_number")
	
	# Get enrolled courses from Academic Analyzer API
	enrolled_courses = []
	try:
		response = requests.get(
			f"{api_base_url()}/student/dashboard",
			params={"rollno": student_roll_number},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get('success'):
				enrolled_courses = [course['courseId'] for course in data.get('courses', [])]
	except requests.RequestException:
		logger.exception("Failed to fetch courses for student active quizzes API")
		return JsonResponse({"error": "Failed to fetch enrolled courses"}, status=500)
	
	# Get today's date for filtering active quizzes
	today = timezone.now()
	
	# Get available quizzes - only for enrolled courses
	if enrolled_courses:
		available_quizzes = Quiz.objects.filter(
			Q(course_id__in=enrolled_courses) & 
			Q(is_active=True) & 
			(Q(start_date__lte=today) | Q(start_date__isnull=True)) & 
			(Q(complete_by_date__gte=today) | Q(complete_by_date__isnull=True))
		).prefetch_related('questions').order_by('-created_at')
	else:
		# If no enrolled courses found, return empty result
		available_quizzes = Quiz.objects.none()
	
	# Convert to a list of dictionaries for JSON response
	quiz_data = []
	for quiz in available_quizzes:
		quiz_data.append({
			"id": quiz.id,
			"title": quiz.title,
			"course_id": quiz.course_id,
			"question_count": quiz.questions.count(),
			"complete_by_date": quiz.complete_by_date.isoformat() if quiz.complete_by_date else None,
			"allow_retake": quiz.allow_retake,
		})
	
	return JsonResponse({"quizzes": quiz_data})


def student_logout(request: HttpRequest) -> HttpResponse:
	"""
	Student logout view that clears the session and redirects to the login page.
	"""
	request.session.flush()
	messages.success(request, "You have been logged out.")
	return redirect("academic_integration:student_login")


def student_course_marks(request: HttpRequest, course_id: str) -> HttpResponse:
	"""
	View for students to see all their marks components details in a specific course with percentages.
	This is the view for http://127.0.0.1:8000/academic_integration/student/course/{course_id}/ URL.
	
	This view can also be accessed by staff members when viewing a student's details by passing
	a 'student' parameter with the student's roll number.
	"""
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	
	# Check if this is being accessed from a staff context with a student parameter
	is_staff_view = False
	staff_email = request.session.get("staff_email")
	student_param = request.GET.get("student")
	
	if staff_email and student_param:
		# Staff is viewing a student's performance
		student_roll_number = student_param
		is_staff_view = True
	else:
		# Regular student access
		student_roll_number = request.session.get("student_roll_number")
		if not student_roll_number:
			messages.info(request, "Please log in to continue.")
			return redirect("academic_integration:student_login")
	
	# Get course details from Academic Analyzer API
	course = {}
	student_marks = {}
	component_details = []
	overall_percentage = 0
	api_error = None
	
	try:
		# First check if the student is enrolled in this course
		dashboard_response = requests.get(
			f"{api_base_url()}/student/dashboard",
			params={"rollno": student_roll_number},
			timeout=5,
		)
		if dashboard_response.ok:
			dashboard_data = _safe_json(dashboard_response)
			if dashboard_data.get('success'):
				courses_data = dashboard_data.get('courses', [])
				enrolled_courses = [c['courseId'] for c in courses_data]
				
				# Find the specific course in the list for basic info
				for c in courses_data:
					if c['courseId'] == course_id:
						course = c
						break
				
				# Check if student is enrolled in this course
				if course_id not in enrolled_courses:
					messages.error(request, "You are not enrolled in this course.")
					return redirect("academic_integration:student_dashboard")
		else:
			logger.warning(f"Failed to fetch student dashboard: {dashboard_response.status_code}")
			api_error = "Failed to fetch course details. Please try again later."
		
		# Now get detailed course marks from the new API endpoint
		marks_response = requests.get(
			f"{api_base_url()}/student/course-marks",
			params={"rollno": student_roll_number, "courseId": course_id},
			timeout=5,
		)
		if marks_response.ok:
			marks_data = _safe_json(marks_response)
			if marks_data.get('success'):
				# Update course info if available
				if marks_data.get('courseName'):
					course['courseName'] = marks_data.get('courseName')
				if marks_data.get('courseCode'):
					course['courseCode'] = marks_data.get('courseCode')
				
				student_marks = marks_data.get('marks', {})
				component_details = marks_data.get('components', [])
				overall_percentage = marks_data.get('overallPercentage', 0)
				
				logger.info(f"Retrieved {len(component_details)} component details for student {student_roll_number} in course {course_id}")
			else:
				logger.warning(f"API returned success=false: {marks_data.get('message')}")
				api_error = marks_data.get('message') or "Failed to fetch course marks details."
		else:
			logger.warning(f"Failed to fetch course marks: {marks_response.status_code}")
			api_error = "Failed to fetch course marks details. Please try again later."
	except requests.RequestException as e:
		logger.exception(f"Error fetching course marks data: {e}")
		api_error = "Could not reach Academic Analyzer API. Please try again later."
	
	# Get quiz performance for this course from our Django database
	from quiz.models import Quiz, QuizAttempt
	
	# Find all quizzes for this course
	quizzes = Quiz.objects.filter(course_id=course_id)
	
	# Get student attempts for these quizzes
	quiz_attempts = []
	if quizzes:
		attempts = QuizAttempt.objects.filter(
			user__username=student_roll_number,
			quiz__in=quizzes,
			completed_at__isnull=False
		).select_related('quiz').order_by('-completed_at')
		
		for attempt in attempts:
			quiz_attempts.append({
				'quiz_title': attempt.quiz.title,
				'score': attempt.score,
				'total_points': attempt.total_points,
				'percentage': attempt.percentage,
				'completed_at': attempt.completed_at
			})
	
	# If we don't have component details but we have quiz attempts,
	# add quiz attempts to our component details
	if not component_details and quiz_attempts:
		total_quiz_score = sum(attempt.get('score', 0) for attempt in quiz_attempts)
		total_quiz_points = sum(attempt.get('total_points', 0) for attempt in quiz_attempts)
		
		if total_quiz_points > 0:
			quiz_percentage = (total_quiz_score / total_quiz_points) * 100
			component_details.append({
				'name': 'Quizzes',
				'type': 'quiz',
				'score': total_quiz_score,
				'maxScore': total_quiz_points,
				'percentage': quiz_percentage,
				'weight': 25  # Default weight
			})
	
	# Calculate grade based on overall percentage
	grade = 'N/A'
	if overall_percentage >= 90:
		grade = 'A'
	elif overall_percentage >= 80:
		grade = 'B'
	elif overall_percentage >= 70:
		grade = 'C'
	elif overall_percentage >= 60:
		grade = 'D'
	elif overall_percentage > 0:
		grade = 'F'
	
	# Serialize the component details and quiz attempts for JavaScript
	component_details_json = json.dumps(component_details, cls=DjangoJSONEncoder)
	quiz_attempts_json = json.dumps(quiz_attempts, cls=DjangoJSONEncoder)
	
	# Add debug info
	debug_info = {
		"api_response": marks_data if 'marks_data' in locals() else None,
		"component_details_count": len(component_details) if component_details else 0,
		"component_details_sample": component_details[0] if component_details and len(component_details) > 0 else None,
	}
	
	context = {
		"student_name": request.session.get("student_name") or student_roll_number,
		"student_roll_number": student_roll_number,
		"course": course,
		"student_marks": student_marks,
		"component_details": component_details,
		"component_details_json": component_details_json,
		"quiz_attempts_json": quiz_attempts_json,
		"overall_percentage": overall_percentage,
		"grade": grade,
		"quiz_attempts": quiz_attempts,
		"api_error": api_error,
		"debug_info": debug_info,
		"is_staff_view": is_staff_view,
		"staff_name": request.session.get("staff_name") if is_staff_view else None,
		"staff_email": staff_email if is_staff_view else None,
	}
	
	return render(request, "academic_integration/student_course_marks.html", context)


def create_course(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to create a new course.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")

	form = CourseForm(request.POST or None)

	if request.method == "POST" and form.is_valid():
		payload = form.cleaned_data
		try:
			response = requests.post(
				f"{api_base_url()}/staff/create-course",
				json={
					"teacherEmail": staff_email,
					"courseName": payload["course_name"],
					"courseCode": payload["course_code"],
					"batch": payload["batch"]
				},
				timeout=5,
			)
		except requests.RequestException:
			logger.exception("Course creation request failed")
			form.add_error(None, "Cannot reach Academic Analyzer API right now. Please try again later.")
		else:
			body = _safe_json(response)
			if response.ok and body.get("success"):
				messages.success(request, body.get("message", "Course created successfully."))
				return redirect("academic_integration:staff_dashboard")
			error_message = body.get("message", "Failed to create course. Please try again.")
			form.add_error(None, error_message)

	context = {
		"form": form,
		"staff_email": staff_email,
		"staff_name": request.session.get("staff_name", staff_email),
	}
	return render(request, "academic_integration/create_course.html", context)


def create_student(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to add a new student to the system.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")

	form = StudentForm(request.POST or None)

	if request.method == "POST" and form.is_valid():
		payload = form.cleaned_data
		try:
			response = requests.post(
				f"{api_base_url()}/staff/create-student",
				json={
					"teacherEmail": staff_email,
					"studentName": payload["name"],
					"rollno": payload["rollno"],
					"batch": payload["batch"],
					"studentEmail": payload["email"],
					"password": payload["password"] or payload["rollno"]  # Use rollno as password if not provided
				},
				timeout=5,
			)
		except requests.RequestException:
			logger.exception("Student creation request failed")
			form.add_error(None, "Cannot reach Academic Analyzer API right now. Please try again later.")
		else:
			body = _safe_json(response)
			if response.ok and body.get("success"):
				messages.success(request, "Student created successfully.")
				return redirect("academic_integration:manage_students")
			error_message = body.get("message", "Failed to create student. Please try again.")
			form.add_error(None, error_message)

	context = {
		"form": form,
		"staff_email": staff_email,
		"staff_name": request.session.get("staff_name", staff_email),
	}
	return render(request, "academic_integration/create_student.html", context)


def create_students_csv(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to bulk add students from a CSV file.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")

	form = CSVUploadForm(request.POST or None, request.FILES or None)

	if request.method == "POST" and form.is_valid():
		csv_file = request.FILES["csv_file"]
		
		# Read the file content
		csv_data = csv_file.read().decode("utf-8")
		
		try:
			response = requests.post(
				f"{api_base_url()}/staff/create-students-csv",
				json={
					"teacherEmail": staff_email,
					"csvData": csv_data
				},
				timeout=10,  # Longer timeout for bulk operations
			)
		except requests.RequestException:
			logger.exception("Bulk student creation request failed")
			form.add_error(None, "Cannot reach Academic Analyzer API right now. Please try again later.")
		else:
			body = _safe_json(response)
			if response.ok and body.get("success"):
				results = body.get("results", {})
				messages.success(
					request, 
					f"Created {results.get('created', 0)} students. "
					f"{results.get('alreadyExists', 0)} already existed. "
					f"{results.get('failed', 0)} failed."
				)
				return redirect("academic_integration:manage_students")
			error_message = body.get("message", "Failed to create students. Please check your CSV format.")
			form.add_error(None, error_message)

	context = {
		"form": form,
		"staff_email": staff_email,
		"staff_name": request.session.get("staff_name", staff_email),
		"csv_template": "rollno,name,batch,email,password\n24MX112,Student Name,24MXG1,student@example.com,password123"
	}
	return render(request, "academic_integration/create_students_csv.html", context)


def manage_students(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to manage students - provides links to different student management actions.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")

	context = {
		"staff_email": staff_email,
		"staff_name": request.session.get("staff_name", staff_email),
	}
	return render(request, "academic_integration/manage_students.html", context)


def view_all_students(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to see all students in the system.
	Supports filtering by name, roll number, batch, or email.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")

	# Get filter parameters
	filter_name = request.GET.get("name", "").strip()
	filter_batch = request.GET.get("batch", "").strip()
	filter_rollno = request.GET.get("rollno", "").strip()
	filter_email = request.GET.get("email", "").strip()
	
	# Track if any filters are applied
	filters_applied = bool(filter_name or filter_batch or filter_rollno or filter_email)

	students = []
	api_error = None

	try:
		# Get all students from the Academic Analyzer API
		params = {"email": staff_email}
		
		# Add filter parameters to the API request if they exist
		if filter_name:
			params["name"] = filter_name
		if filter_batch:
			params["batch"] = filter_batch
		if filter_rollno:
			params["rollno"] = filter_rollno
		if filter_email:
			params["student_email"] = filter_email
			
		response = requests.get(
			f"{api_base_url()}/staff/all-students",
			params=params,
			timeout=10,
		)
	except requests.RequestException as e:
		logger.exception(f"Failed to load student data: {str(e)}")
		api_error = "Could not reach Academic Analyzer API. Please check your internet connection and refresh the page."
	else:
		body = _safe_json(response)
		if response.ok and body.get("success"):
			students = body.get("students", [])
			logger.info(f"Successfully loaded {len(students)} students with filters applied: {filters_applied}")
			
			# If API filtering is not supported, apply filters in Python
			if filters_applied and not any(param in params for param in ["name", "batch", "rollno", "student_email"]):
				filtered_students = []
				for student in students:
					include_student = True
					
					if filter_name and filter_name.lower() not in student.get("name", "").lower():
						include_student = False
					
					if filter_batch and filter_batch.lower() not in student.get("batch", "").lower():
						include_student = False
						
					if filter_rollno and filter_rollno.lower() not in student.get("rollno", "").lower():
						include_student = False
						
					if filter_email and filter_email.lower() not in student.get("email", "").lower():
						include_student = False
						
					if include_student:
						filtered_students.append(student)
						
				students = filtered_students
				logger.info(f"Applied filters in Python, resulting in {len(students)} students")
		else:
			error_message = body.get("message", "Unknown error")
			logger.error(f"API Error in view_all_students: {error_message}")
			api_error = f"API Error: {error_message}. Please try again later."

	# Get unique batches for the batch dropdown filter
	batches = sorted(set(student.get("batch", "") for student in students if student.get("batch")))

	context = {
		"staff_name": request.session.get("staff_name") or staff_email,
		"staff_email": staff_email,
		"students": students,
		"api_error": api_error,
		"filters": {
			"name": filter_name,
			"batch": filter_batch,
			"rollno": filter_rollno,
			"email": filter_email
		},
		"filters_applied": filters_applied,
		"batches": batches,
		"total_students": len(students)
	}
	return render(request, "academic_integration/view_all_students.html", context)


def student_detail(request: HttpRequest, rollno: str) -> HttpResponse:
	"""
	View for staff to see detailed information about a specific student.
	"""
	from quiz.models import QuizAttempt
	from django.db.models import Avg
	
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")

	student = None
	enrolled_courses = []
	api_error = None

	try:
		# Get student details from the Academic Analyzer API
		response = requests.get(
			f"{api_base_url()}/staff/student-detail",
			params={"email": staff_email, "rollno": rollno},
			timeout=10,
		)
	except requests.RequestException as e:
		logger.exception(f"Failed to load student detail: {str(e)}")
		api_error = "Could not reach Academic Analyzer API. Please check your internet connection and refresh the page."
	else:
		body = _safe_json(response)
		if response.ok and body.get("success"):
			student = body.get("student", {})
			enrolled_courses = body.get("courses", [])
			logger.info(f"Successfully loaded details for student {rollno}")
		else:
			error_message = body.get("message", "Unknown error")
			logger.error(f"API Error in student_detail: {error_message}")
			api_error = f"API Error: {error_message}. Please try again later."

	# Get quiz attempts from the local database
	quiz_attempts = []
	completed_quizzes = 0
	avg_score = 0
	
	if student:
		quiz_attempts = QuizAttempt.objects.filter(
			user__username=rollno,
			completed_at__isnull=False
		).select_related('quiz').order_by('-completed_at')[:10]  # Get last 10 attempts
		
		# Add course names to quiz attempts
		course_lookup = {course['courseId']: course['courseName'] for course in enrolled_courses}
		for attempt in quiz_attempts:
			attempt.course_name = course_lookup.get(attempt.quiz.course_id, "Unknown Course")
		
		# Calculate statistics
		completed_quizzes = quiz_attempts.count()
		avg_score_obj = quiz_attempts.aggregate(Avg('percentage'))
		avg_score = avg_score_obj['percentage__avg'] or 0

	context = {
		"staff_name": request.session.get("staff_name") or staff_email,
		"staff_email": staff_email,
		"student": student,
		"enrolled_courses": enrolled_courses,
		"quiz_attempts": quiz_attempts,
		"completed_quizzes": completed_quizzes,
		"avg_score": avg_score,
		"api_error": api_error,
	}
	return render(request, "academic_integration/student_detail.html", context)


def manage_course(request: HttpRequest, course_id: str) -> HttpResponse:
	"""
	View for staff to manage a specific course - view roster, add students, analytics, etc.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")

	# Get course details
	api_error = None
	course = {}
	students = []
	sorted_students = []
	
	# Fetch available batches for batch enrollment form
	batches = []
	try:
		batch_response = requests.get(
			f"{api_base_url()}/staff/all-batches",
			timeout=5,
		)
		if batch_response.ok:
			batch_body = _safe_json(batch_response)
			if batch_body.get("success"):
				batches = batch_body.get("batches", [])
	except requests.RequestException:
		logger.warning("Failed to fetch batches from API")

	try:
		response = requests.get(
			f"{api_base_url()}/staff/course-detail",
			params={"courseId": course_id},
			timeout=5,
		)
	except requests.RequestException:
		logger.exception("Failed to load course details")
		api_error = "Could not reach Academic Analyzer API. Please try again later."
	else:
		body = _safe_json(response)
		if response.ok and body.get("success"):
			course = {
				"id": course_id, 
				"courseId": course_id,
				"name": body.get("courseName", "Unknown Course"),
				"courseName": body.get("courseName", "Unknown Course"),
				"courseCode": body.get("courseCode", ""),
				"batch": body.get("batch", "")
			}
			students = body.get("students", [])
			# Sort students by roll number for display
			sorted_students = sorted(students, key=lambda x: x.get('rollno', ''))
		else:
			api_error = body.get("message", "Failed to load course details.")

	# Forms for adding students
	single_student_form = StudentAddForm(request.POST or None if request.POST.get("form_type") == "single" else None)
	batch_form = BatchEnrollmentForm(
		request.POST or None if request.POST.get("form_type") == "batch" else None,
		batches=batches
	)
	csv_form = CSVUploadForm(request.POST or None, request.FILES or None if request.POST.get("form_type") == "csv" else None)

	# Process single student form
	if request.method == "POST" and request.POST.get("form_type") == "single" and single_student_form.is_valid():
		payload = single_student_form.cleaned_data
		rollno = payload["rollno"].strip().upper()  # Make case-insensitive by converting to uppercase
		
		logger.info(f"Attempting to add student with rollno: {rollno} to course: {course_id}")
		
		try:
			# Try with the rollno directly as the identifier
			api_payload = {
				"teacherEmail": staff_email,
				"courseId": course_id,
				"rollno": rollno  # Use rollno instead of studentEmail
			}
			
			logger.info(f"Sending request to Academic Analyzer API: {api_payload}")
			
			response = requests.post(
				f"{api_base_url()}/staff/add-student",
				json=api_payload,
				timeout=5,
			)
			
			logger.info(f"API Response Status: {response.status_code}, Body: {response.text}")
			
		except requests.RequestException as e:
			logger.exception(f"API request failed: {str(e)}")
			single_student_form.add_error(None, "Cannot reach Academic Analyzer API. Please try again later.")
		else:
			body = _safe_json(response)
			if response.ok and body.get("success"):
				messages.success(request, body.get("message", "Student added successfully."))
				return redirect("academic_integration:manage_course", course_id=course_id)
			else:
				error_msg = body.get("message", "Failed to add student.")
				logger.warning(f"Failed to add student {rollno}: {error_msg}")
				single_student_form.add_error(None, f"{error_msg} (Roll No: {rollno})")

	# Process batch enrollment form
	if request.method == "POST" and request.POST.get("form_type") == "batch" and batch_form.is_valid():
		payload = batch_form.cleaned_data
		try:
			response = requests.post(
				f"{api_base_url()}/staff/add-batch-to-course",
				json={
					"teacherEmail": staff_email,
					"courseId": course_id,
					"batch": payload["batch"]
				},
				timeout=10,  # Longer timeout for batch operations
			)
		except requests.RequestException:
			batch_form.add_error(None, "Cannot reach Academic Analyzer API. Please try again later.")
		else:
			body = _safe_json(response)
			if response.ok and body.get("success"):
				results = body.get("results", {})
				messages.success(
					request,
					f"Added {results.get('added', 0)} students. "
					f"{results.get('alreadyEnrolled', 0)} were already enrolled."
				)
				return redirect("academic_integration:manage_course", course_id=course_id)
			batch_form.add_error(None, body.get("message", "Failed to enroll batch."))

	# Process CSV/Excel upload form
	if request.method == "POST" and request.POST.get("form_type") == "csv" and csv_form.is_valid():
		import csv
		import io
		
		upload_file = request.FILES["csv_file"]
		file_name = upload_file.name.lower()
		
		# Extract roll numbers from file
		roll_numbers = []
		
		try:
			if file_name.endswith('.csv'):
				# Read CSV file
				csv_data = upload_file.read().decode("utf-8")
				csv_reader = csv.reader(io.StringIO(csv_data))
				for row in csv_reader:
					if row and row[0].strip():  # Skip empty rows
						roll_numbers.append(row[0].strip())
			elif file_name.endswith(('.xlsx', '.xls')):
				# Read Excel file
				try:
					import openpyxl
					import xlrd
				except ImportError:
					csv_form.add_error(None, "Excel support not installed. Please install openpyxl and xlrd packages.")
				else:
					if file_name.endswith('.xlsx'):
						# Read .xlsx with openpyxl
						workbook = openpyxl.load_workbook(upload_file)
						sheet = workbook.active
						for row in sheet.iter_rows(min_row=1, values_only=True):
							if row and row[0]:
								roll_numbers.append(str(row[0]).strip())
					else:
						# Read .xls with xlrd
						workbook = xlrd.open_workbook(file_contents=upload_file.read())
						sheet = workbook.sheet_by_index(0)
						for row_idx in range(sheet.nrows):
							cell_value = sheet.cell_value(row_idx, 0)
							if cell_value:
								roll_numbers.append(str(cell_value).strip())
			else:
				csv_form.add_error(None, "Invalid file format. Please upload CSV or Excel file.")
				roll_numbers = []
			
			# Convert roll numbers to CSV format for API
			csv_data = '\n'.join(roll_numbers)
			
		except Exception as e:
			logger.exception(f"Error reading file: {e}")
			csv_form.add_error(None, f"Error reading file: {str(e)}")
			csv_data = None
		
		# Send to API if we have data
		if csv_data:
			# No need to normalize here - the API now handles case-insensitive matching
			logger.info(f"Processing student list upload for course: {course_id}")
			
			try:
				response = requests.post(
					f"{api_base_url()}/staff/add-students-csv",
					json={
						"teacherEmail": staff_email,
						"courseId": course_id,
						"csvData": csv_data
					},
					timeout=10,
				)
				
				logger.info(f"Student list Upload API Response Status: {response.status_code}, Body: {response.text}")
				
			except requests.RequestException as e:
				logger.exception(f"Student list API request failed: {str(e)}")
				csv_form.add_error(None, "Cannot reach Academic Analyzer API. Please try again later.")
			else:
				body = _safe_json(response)
				if response.ok and body.get("success"):
					results = body.get("results", {})
					messages.success(
						request,
						f"Added {results.get('added', 0)} students. "
						f"{results.get('notFound', 0)} not found. "
						f"{results.get('alreadyEnrolled', 0)} already enrolled."
					)
					return redirect("academic_integration:manage_course", course_id=course_id)
				else:
					error_msg = body.get("message", "Failed to process file.")
					logger.warning(f"Student list upload failed: {error_msg}")
				csv_form.add_error(None, error_msg)

	# Process bulk marks upload
	bulk_marks_errors = []
	bulk_marks_success = None
	
	# Initialize direct marks variables
	direct_marks_errors = []
	direct_marks_success = None
	
	# Log POST data for debugging
	if request.method == "POST":
		logger.info(f"POST request received - form_type: {request.POST.get('form_type')}")
		logger.info(f"Files in request: {list(request.FILES.keys())}")
	
	if request.method == "POST" and request.POST.get("form_type") == "bulk_marks":
		logger.info("Processing bulk marks upload")
		if 'marks_csv_file' in request.FILES:
			import csv
			import io
			import re
			
			marks_file = request.FILES['marks_csv_file']
			file_name = marks_file.name.lower()
			
			try:
				# Determine file type and read accordingly
				if file_name.endswith('.csv'):
					# Read CSV file
					csv_data = marks_file.read().decode('utf-8')
					csv_reader = csv.DictReader(io.StringIO(csv_data))
					rows = list(csv_reader)
				elif file_name.endswith(('.xlsx', '.xls')):
					# Read Excel file
					try:
						import openpyxl
						import xlrd
					except ImportError:
						bulk_marks_errors.append("Excel support not installed. Please install openpyxl and xlrd packages.")
						rows = []
					else:
						if file_name.endswith('.xlsx'):
							# Read .xlsx with openpyxl
							workbook = openpyxl.load_workbook(marks_file)
							sheet = workbook.active
							# Convert to list of dicts
							headers = [cell.value for cell in sheet[1]]
							rows = []
							for row in sheet.iter_rows(min_row=2, values_only=True):
								row_dict = dict(zip(headers, row))
								rows.append(row_dict)
						else:
							# Read .xls with xlrd
							workbook = xlrd.open_workbook(file_contents=marks_file.read())
							sheet = workbook.sheet_by_index(0)
							# Convert to list of dicts
							headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
							rows = []
							for row_idx in range(1, sheet.nrows):
								row_dict = dict(zip(headers, [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]))
								rows.append(row_dict)
				else:
					bulk_marks_errors.append("Invalid file format. Please upload CSV or Excel file.")
					rows = []
				
				# Column name mapping - flexible matching
				def normalize_column_name(col):
					"""Normalize column names for flexible matching"""
					if col is None:
						return ''
					col = str(col).lower().strip()
					col = re.sub(r'[^a-z0-9]', '', col)  # Remove special chars
					return col
				
				# Map common column names to API field names
				column_mapping = {
					'tutorial1': 'tutorial1',
					'tutorial2': 'tutorial2',
					'tutorial3': 'tutorial3',
					'tutorial4': 'tutorial4',
					'ca1': 'CA1',
					'ca2': 'CA2',
					'assignment': 'assignmentPresentation',
					'presentation': 'assignmentPresentation',
					'assignmentpresentation': 'assignmentPresentation',
				}
				
				# Track updates
				updates_count = 0
				errors_count = 0
				
				# Process each row
				for row_num, row in enumerate(rows, start=2):  # Start at 2 (header is row 1)
					# Get student identifier (roll number or email)
					student_identifier = str(row.get('Roll Number', '') or row.get('rollno', '') or '').strip()
					student_email = str(row.get('Email', '') or row.get('email', '') or '').strip()
					
					if not student_identifier and not student_email:
						bulk_marks_errors.append(f"Row {row_num}: Missing student identifier")
						errors_count += 1
						continue
					
					# Collect marks to update (only non-empty values)
					marks_updates = {}
					
					for col_name, value in row.items():
						if value is None or str(value).strip() == '':
							continue  # Skip empty values
						
						# Normalize column name
						normalized = normalize_column_name(col_name)
						
						# Check if it matches a marks column
						if normalized in column_mapping:
							api_field = column_mapping[normalized]
							try:
								mark_value = float(str(value).strip())
								if 0 <= mark_value <= 10:
									marks_updates[api_field] = mark_value
								else:
									bulk_marks_errors.append(f"Row {row_num}, {col_name}: Mark {mark_value} out of range (0-10)")
									errors_count += 1
							except (ValueError, TypeError):
								bulk_marks_errors.append(f"Row {row_num}, {col_name}: Invalid number '{value}'")
								errors_count += 1
					
					# If we have marks to update, send to API
					if marks_updates:
						# Prepare API request for each mark type
						for api_field, mark_value in marks_updates.items():
							# Map field to API endpoint
							endpoint_mapping = {
								'tutorial1': 'add-tut1-mark',
								'tutorial2': 'add-tut2-mark',
								'tutorial3': 'add-tut3-mark',
								'tutorial4': 'add-tut4-mark',
								'CA1': 'add-ca1-mark',
								'CA2': 'add-ca2-mark',
								'assignmentPresentation': 'add-assignment-mark'
							}
							
							endpoint = endpoint_mapping.get(api_field)
							if not endpoint:
								continue
							
							try:
								# Send update to API
								# Prepare student input - prefer rollno if available
								student_input = {"mark": mark_value}
								if student_identifier:
									student_input["rollno"] = student_identifier
								if student_email:
									student_input["email"] = student_email
								
								api_response = requests.post(
									f"{api_base_url()}/staff/{endpoint}",
									json={
										"teacherEmail": staff_email,
										"courseId": course_id,
										"studentInput": [student_input]
									},
									timeout=5,
								)
								
								if api_response.ok and _safe_json(api_response).get("success"):
									updates_count += 1
								else:
									error_msg = _safe_json(api_response).get("message", "Unknown error")
									bulk_marks_errors.append(f"Row {row_num}: {error_msg}")
									errors_count += 1
									
							except requests.RequestException as e:
								bulk_marks_errors.append(f"Row {row_num}: API connection error")
								errors_count += 1
								logger.exception(f"Bulk marks API error: {e}")
				
				# Set success message
				if updates_count > 0:
					bulk_marks_success = f"Successfully updated {updates_count} mark entries."
					if errors_count > 0:
						bulk_marks_success += f" {errors_count} errors occurred."
				elif errors_count > 0:
					bulk_marks_errors.insert(0, "No marks were updated due to errors.")
				else:
					bulk_marks_errors.append("No valid marks found in the CSV file.")
					
			except Exception as e:
				logger.exception(f"Error processing bulk marks CSV: {e}")
				bulk_marks_errors.append(f"File processing error: {str(e)}")
		else:
			bulk_marks_errors.append("No file uploaded")

	# Process direct mark entry
	if request.method == "POST" and request.POST.get("form_type") == "direct_marks":
		mark_component = request.POST.get("mark_component")
		
		logger.info(f"Direct marks submission - Component: {mark_component}")
		logger.info(f"POST data keys: {list(request.POST.keys())}")
		
		if not mark_component:
			direct_marks_errors.append("Please select a mark component")
		else:
			# Get total marks for conversion
			total_marks = request.POST.get("total_marks")
			try:
				total_marks = float(total_marks) if total_marks else 10.0
				if total_marks <= 0:
					total_marks = 10.0
			except ValueError:
				total_marks = 10.0
			
			logger.info(f"Total marks for test: {total_marks}")
			
			# Map component to API endpoint
			endpoint_mapping = {
				'tutorial1': 'add-tut1-mark',
				'tutorial2': 'add-tut2-mark',
				'tutorial3': 'add-tut3-mark',
				'tutorial4': 'add-tut4-mark',
				'ca1': 'add-ca1-mark',
				'ca2': 'add-ca2-mark',
				'assignment': 'add-assignment-mark'
			}
			
			endpoint = endpoint_mapping.get(mark_component)
			if not endpoint:
				direct_marks_errors.append(f"Invalid component: {mark_component}")
			else:
				# Collect all marks from the form
				updates_count = 0
				errors_count = 0
				student_inputs = []
				
				# Get all student rollnos and their marks
				counter = 1
				while f"student_rollno_{counter}" in request.POST:
					rollno = request.POST.get(f"student_rollno_{counter}")
					email = request.POST.get(f"student_email_{counter}")
					mark_value = request.POST.get(f"mark_{counter}")
					
					if mark_value and mark_value.strip():
						try:
							actual_mark = float(mark_value.strip())
							
							# Validate against total marks
							if 0 <= actual_mark <= total_marks:
								# Convert to equivalent out of 10
								equivalent_mark = (actual_mark / total_marks) * 10
								# Round to 2 decimal places
								equivalent_mark = round(equivalent_mark, 2)
								
								student_input = {
									"rollno": rollno,
									"mark": equivalent_mark
								}
								if email:
									student_input["email"] = email
								student_inputs.append(student_input)
								
								logger.info(f"{rollno}: {actual_mark}/{total_marks} = {equivalent_mark}/10")
							else:
								direct_marks_errors.append(f"{rollno}: Mark {actual_mark} out of range (0-{total_marks})")
								errors_count += 1
						except ValueError:
							direct_marks_errors.append(f"{rollno}: Invalid mark value '{mark_value}'")
							errors_count += 1
					
					counter += 1
				
				logger.info(f"Collected {len(student_inputs)} student marks: {student_inputs}")
				
				# Send to API if we have marks to update
				if student_inputs:
					try:
						api_url = f"{api_base_url()}/staff/{endpoint}"
						api_payload = {
							"teacherEmail": staff_email,
							"courseId": course_id,
							"studentInput": student_inputs
						}
						logger.info(f"Sending to API: {api_url}")
						logger.info(f"Payload: {api_payload}")
						
						api_response = requests.post(
							api_url,
							json=api_payload,
							timeout=10,
						)
						
						logger.info(f"API response status: {api_response.status_code}")
						logger.info(f"API response body: {_safe_json(api_response)}")
						
						if api_response.ok and _safe_json(api_response).get("success"):
							updates_count = len(student_inputs)
							direct_marks_success = f"Successfully updated {updates_count} student marks for {mark_component.upper()}."
						else:
							error_msg = _safe_json(api_response).get("message", "Unknown error")
							direct_marks_errors.append(f"API Error: {error_msg}")
							errors_count += 1
							
					except requests.RequestException as e:
						direct_marks_errors.append("API connection error. Please try again.")
						errors_count += 1
						logger.exception(f"Direct marks API error: {e}")
				else:
					direct_marks_errors.append("No marks entered. Please enter at least one mark.")

	# Get performance statistics for the course using our new endpoint
	overall_stats = {}
	# Only fetch analytics if we successfully retrieved the course details
	if not api_error and students:
		try:
			logger.info(f"Fetching analytics data for course: {course_id}")
			response = requests.get(
				f"{api_base_url()}/staff/course-analytics",
				params={"courseId": course_id},
				timeout=15,  # Increased timeout for analytics data which might be complex
			)
			if response.ok:
				data = _safe_json(response)
				if data.get("success"):
					# Get overall stats directly from the API
					overall_stats = data.get("overallStats", {})
					logger.info(f"Successfully loaded analytics data for course: {course_id}")
					
					# Calculate grade distribution percentages for progress bars
					if "gradeDistribution" in overall_stats:
						dist = overall_stats["gradeDistribution"]
						total_students = len(students) or 1  # Avoid division by zero
						
						# Calculate percentage for each grade (for progress bars)
						overall_stats["grade_a_percentage"] = (dist.get("A", 0) / total_students) * 100 if total_students > 0 else 0
						overall_stats["grade_b_percentage"] = (dist.get("B", 0) / total_students) * 100 if total_students > 0 else 0
						overall_stats["grade_c_percentage"] = (dist.get("C", 0) / total_students) * 100 if total_students > 0 else 0
						overall_stats["grade_d_percentage"] = (dist.get("D", 0) / total_students) * 100 if total_students > 0 else 0
						overall_stats["grade_e_percentage"] = (dist.get("E", 0) / total_students) * 100 if total_students > 0 else 0
						overall_stats["grade_f_percentage"] = (dist.get("F", 0) / total_students) * 100 if total_students > 0 else 0
					
					# Calculate score range distribution for charts
					score_ranges = {
						"score_range_0_20": 0,
						"score_range_21_40": 0,
						"score_range_41_60": 0,
						"score_range_61_70": 0,
						"score_range_71_80": 0,
						"score_range_81_90": 0,
						"score_range_91_100": 0,
					}
					
					# Get detailed student performances from the API
					student_performances = data.get("studentPerformances", {})
					
					# Count students in each score range based on finalInternal score
					for rollno, perf in student_performances.items():
						score = perf.get("finalInternal", 0)
						percentage = (score / 50) * 100  # Convert to percentage (score is out of 50)
						
						if percentage < 20:
							score_ranges["score_range_0_20"] += 1
						elif percentage < 40:
							score_ranges["score_range_21_40"] += 1
						elif percentage < 60:
							score_ranges["score_range_41_60"] += 1
						elif percentage < 70:
							score_ranges["score_range_61_70"] += 1
						elif percentage < 80:
							score_ranges["score_range_71_80"] += 1
						elif percentage < 90:
							score_ranges["score_range_81_90"] += 1
						else:
							score_ranges["score_range_91_100"] += 1
					
					# Add score ranges to overall stats
					overall_stats.update(score_ranges)
					
					# Calculate component-wise performance percentages for radar chart
					tutorials_max = 15  # Max points for tutorials
					cas_max = 20       # Max points for CAs
					assignment_max = 15 # Max points for assignments
					
					# Get component averages directly from the API
					component_avgs = overall_stats.get("componentAverages", {})
					
					# Calculate percentages for radar chart
					tut_avg = (
						component_avgs.get("tutorial1", 0) + 
						component_avgs.get("tutorial2", 0) + 
						component_avgs.get("tutorial3", 0) + 
						component_avgs.get("tutorial4", 0)
					) / 4  # Average of all tutorials
					
					ca_avg = (component_avgs.get("CA1", 0) + component_avgs.get("CA2", 0)) / 2
					assignment_avg = component_avgs.get("assignmentPresentation", 0)
					
					# Convert to percentages for chart display
					overall_stats["tutorial_percentage"] = (tut_avg / 10) * 100  # Tutorial scores out of 10
					overall_stats["ca_percentage"] = (ca_avg / 10) * 100  # CA scores out of 10
					overall_stats["assignment_percentage"] = (assignment_avg / 15) * 100  # Assignment out of 15
					
					# Add detailed performance data to each student
					for student in students:
						roll_number = student.get("rollno")
						if roll_number in student_performances:
							student.update(student_performances[roll_number])
				else:
					api_error = api_error or data.get("message", "Failed to load performance data.")
					logger.error(f"API returned error: {data.get('message', 'Unknown error')}")
			else:
				error_message = data.get("message", "Unknown error")
				logger.error(f"API failed to load performance data: {error_message}")
				api_error = api_error or f"API error: {error_message}"
		except requests.RequestException as e:
			logger.exception(f"Failed to load course analytics performance data: {str(e)}")
			api_error = api_error or f"Could not reach Academic Analyzer API for performance data: {str(e)}. Please check your network connection and try again."
		except Exception as e:
			logger.exception(f"Unexpected error processing analytics data: {str(e)}")
			api_error = api_error or f"An unexpected error occurred while processing analytics data: {str(e)}. Please try refreshing the page."
	
	# Get quiz results for the course
	from quiz.models import Quiz, QuizAttempt
	from django.db.models import Avg
	
	quizzes = []
	quiz_stats = []
	
	try:
		quizzes = Quiz.objects.filter(course_id=course_id).order_by('-created_at')
		
		for quiz in quizzes:
			attempts = QuizAttempt.objects.filter(quiz=quiz, completed_at__isnull=False)
			stats = {
				"quiz_id": quiz.id,
				"title": quiz.title,
				"attempt_count": attempts.count(),
				"avg_score": attempts.aggregate(Avg('percentage'))['percentage__avg'] or 0,
			}
			quiz_stats.append(stats)
	except Exception as e:
		logger.exception(f"Error retrieving quiz data: {str(e)}")
		# Don't fail the entire request if just the quiz data fails
		if not api_error:
			api_error = "Warning: Could not load quiz statistics."

	context = {
		"staff_email": staff_email,
		"staff_name": request.session.get("staff_name", staff_email),
		"course": course,
		"students": students,
		"sorted_students": sorted_students,
		"api_error": api_error,
		"single_student_form": single_student_form,
		"batch_form": batch_form,
		"csv_form": csv_form,
		"csv_template": "rollno\n24MX112\n24MX113",
		# Analytics data
		"overall_stats": overall_stats,
		"quiz_stats": quiz_stats,
		"tutorial_max_marks": 10,  # Default max marks for tutorials
		# Bulk marks upload
		"bulk_marks_errors": bulk_marks_errors,
		"bulk_marks_success": bulk_marks_success,
		# Direct marks entry
		"direct_marks_errors": direct_marks_errors,
		"direct_marks_success": direct_marks_success,
	}
	return render(request, "academic_integration/manage_course.html", context)


def remove_student_from_course(request: HttpRequest, course_id: str) -> HttpResponse:
	"""
	View to remove a student from a course.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.error(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")
	
	if request.method != "POST":
		messages.error(request, "Invalid request method.")
		return redirect("academic_integration:manage_course", course_id=course_id)
	
	student_rollno = request.POST.get("student_rollno")
	
	if not student_rollno:
		messages.error(request, "Student roll number is required.")
		return redirect("academic_integration:manage_course", course_id=course_id)
	
	try:
		response = requests.post(
			f"{api_base_url()}/staff/remove-student",
			json={
				"teacherEmail": staff_email,
				"courseId": course_id,
				"studentRollno": student_rollno
			},
			timeout=5,
		)
		
		body = _safe_json(response)
		
		if response.ok and body.get("success"):
			messages.success(request, body.get("message", "Student removed successfully."))
		else:
			messages.error(request, body.get("message", "Failed to remove student."))
			
	except requests.RequestException as e:
		logger.exception(f"Failed to remove student: {str(e)}")
		messages.error(request, "Cannot reach Academic Analyzer API. Please try again later.")
	
	return redirect("academic_integration:manage_course", course_id=course_id)


def staff_analytics(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to access analytics and performance data for a specific course.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")
	
	course_id = request.GET.get("course_id")
	if not course_id:
		messages.warning(request, "Please select a course to view analytics.")
		return redirect("academic_integration:staff_dashboard")
	
	# Get course details and student performance
	api_error = None
	course = {}
	students = []
	
	try:
		# First, get course details
		response = requests.get(
			f"{api_base_url()}/staff/course-detail",
			params={"courseId": course_id},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get("success"):
				course = {
					"courseId": course_id,
					"courseName": data.get("courseName", "Unknown Course"),
					"courseCode": data.get("courseCode", ""),
					"batch": data.get("batch", "")
				}
				students = data.get("students", [])
			else:
				api_error = data.get("message", "Failed to load course details.")
		else:
			api_error = "API error: Failed to load course details."
	except requests.RequestException:
		logger.exception("Failed to load course analytics data")
		api_error = "Could not reach Academic Analyzer API. Please try again later."
	
	# We no longer need this function as analytics are fully integrated into the course management page
	# Redirect to manage_course which now has all analytics functionality
	return redirect('academic_integration:manage_course', course_id=course_id)
	
	# Get quiz results for the course
	from quiz.models import Quiz, QuizAttempt
	
	quizzes = Quiz.objects.filter(course_id=course_id).order_by('-created_at')
	quiz_stats = []
	
	for quiz in quizzes:
		attempts = QuizAttempt.objects.filter(quiz=quiz, completed_at__isnull=False)
		stats = {
			"quiz_id": quiz.id,
			"title": quiz.title,
			"attempt_count": attempts.count(),
			"avg_score": attempts.aggregate(Avg('percentage'))['percentage__avg'] or 0,
		}
		quiz_stats.append(stats)
	
	context = {
		"staff_email": staff_email,
		"staff_name": request.session.get("staff_name", staff_email),
		"course": course,
		"students": students,
		"overall_stats": overall_stats,
		"quiz_stats": quiz_stats,
		"api_error": api_error,
	}
	return render(request, "academic_integration/staff_analytics.html", context)


def edit_student_marks(request: HttpRequest) -> HttpResponse:
	"""
	View for staff to edit marks for a student in a specific course.
	"""
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")
	
	# Get parameters
	student_id = request.GET.get("student_id") or request.POST.get("student_id")
	course_id = request.GET.get("course_id") or request.POST.get("course_id")
	
	if not student_id or not course_id:
		messages.warning(request, "Both student and course must be specified.")
		return redirect("academic_integration:staff_dashboard")
	
	# Get course and student details
	api_error = None
	course = {}
	student = {}
	tutorial_max_marks = 10  # Default max marks for tutorials
	
	try:
		# Get course details
		response = requests.get(
			f"{api_base_url()}/staff/course-detail",
			params={"courseId": course_id},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get("success"):
				course = {
					"courseId": course_id,
					"courseName": data.get("courseName", "Unknown Course"),
					"courseCode": data.get("courseCode", "")
				}
			else:
				api_error = data.get("message", "Failed to load course details.")
		else:
			api_error = "API error: Failed to load course details."
		
		# Get student details
		response = requests.get(
			f"{api_base_url()}/staff/student-detail",
			params={"studentId": student_id},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get("success"):
				student = data.get("student", {})
			else:
				api_error = api_error or data.get("message", "Failed to load student details.")
		else:
			api_error = api_error or "API error: Failed to load student details."
			
		# Get performance data for this student in this course using our new API
		response = requests.get(
			f"{api_base_url()}/staff/student-performance",
			params={"studentId": student_id, "courseId": course_id},
			timeout=10,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get("success"):
				# Extract student details
				student = data.get("student", {})
				# Extract performance data
				performance = data.get("performance", {})
				student.update(performance)
				# Get tutorial max marks
				tutorial_max_marks = data.get("tutorialMaxMarks", 10)
				
				logger.info(f"Successfully loaded performance data for student: {student.get('name')}")
			else:
				error_message = data.get("message", "Unknown error")
				logger.error(f"API returned error loading student performance: {error_message}")
				api_error = api_error or data.get("message", "Failed to load performance data.")
		else:
			error_message = data.get("message", "Unknown error")
			logger.error(f"API failed to load student performance data: {error_message}")
			api_error = api_error or f"API error: Failed to load performance data. Status code: {response.status_code}"
	except requests.RequestException:
		logger.exception("Failed to load data for student marks editing")
		api_error = "Could not reach Academic Analyzer API. Please try again later."
	
	# Process form submission
	if request.method == "POST":
		try:
			# Extract form data
			tutorial1 = request.POST.get("tutorial1") or None
			tutorial2 = request.POST.get("tutorial2") or None
			tutorial3 = request.POST.get("tutorial3") or None
			tutorial4 = request.POST.get("tutorial4") or None
			ca1 = request.POST.get("ca1") or None
			ca2 = request.POST.get("ca2") or None
			assignment = request.POST.get("assignment") or None
			presentation = request.POST.get("presentation") or None
			
			# Prepare data for API
			update_data = {
				"studentId": student_id,
				"courseId": course_id,
				"teacherEmail": staff_email,
				"marks": {}
			}
			
			# Only include fields that were filled
			if tutorial1 is not None:
				update_data["marks"]["tutorial1"] = float(tutorial1)
			if tutorial2 is not None:
				update_data["marks"]["tutorial2"] = float(tutorial2)
			if tutorial3 is not None:
				update_data["marks"]["tutorial3"] = float(tutorial3)
			if tutorial4 is not None:
				update_data["marks"]["tutorial4"] = float(tutorial4)
			if ca1 is not None:
				update_data["marks"]["ca1"] = float(ca1)
			if ca2 is not None:
				update_data["marks"]["ca2"] = float(ca2)
			if assignment is not None:
				update_data["marks"]["assignment"] = float(assignment)
			if presentation is not None:
				update_data["marks"]["presentation"] = float(presentation)
			
			# Use our new API endpoint for updating marks
			logger.info(f"Updating marks for student ID {student_id} in course {course_id}")
			response = requests.post(
				f"{api_base_url()}/staff/update-student-marks",
				json=update_data,
				timeout=10,  # Increased timeout for update operations
			)
			
			if response.ok:
				data = _safe_json(response)
				if data.get("success"):
					messages.success(request, "Student marks updated successfully.")
					return redirect(f"{reverse('academic_integration:staff_analytics')}?course_id={course_id}")
				else:
					messages.error(request, data.get("message", "Failed to update marks."))
			else:
				messages.error(request, "API error: Failed to update marks.")
		except ValueError:
			messages.error(request, "Invalid values for marks. Please enter valid numbers.")
		except requests.RequestException:
			logger.exception("Failed to update student marks")
			messages.error(request, "Could not reach Academic Analyzer API. Please try again later.")
	
	context = {
		"staff_email": staff_email,
		"staff_name": request.session.get("staff_name", staff_email),
		"course": course,
		"student": student,
		"tutorial_max_marks": tutorial_max_marks,
		"api_error": api_error,
	}
	return render(request, "academic_integration/edit_student_marks.html", context)


def student_profile(request: HttpRequest) -> HttpResponse:
	"""
	Student profile view that allows students to view and edit their profile information.
	"""
	import logging
	
	# Set up logging
	logger = logging.getLogger(__name__)
	
	# Ensure student is logged in
	student_roll_number = request.session.get("student_roll_number")
	if not student_roll_number:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:student_login")
	
	student_name = request.session.get("student_name", "")
	student_email = request.session.get("student_email", f"{student_roll_number}@student.edu")
	api_error = None
	allow_name_edit = False  # Default to not allowing name editing
	email_notifications = request.session.get("email_notifications", True)  # Default to enabled
	
	# Get student details from Academic Analyzer API
	try:
		response = requests.get(
			f"{api_base_url()}/student/profile",
			params={"rollno": student_roll_number},
			timeout=5,
		)
		if response.ok:
			data = _safe_json(response)
			if data.get('success'):
				# If API returns student profile data, update the session
				student_data = data.get('student', {})
				if student_data.get('name'):
					student_name = student_data.get('name')
					request.session["student_name"] = student_name
				if student_data.get('email'):
					student_email = student_data.get('email')
					request.session["student_email"] = student_email
				
				# Allow name edit only if API specifically allows it
				allow_name_edit = student_data.get('allow_name_edit', False)
				
				# Check if student has email notification preferences
				email_notifications = student_data.get('email_notifications', email_notifications)
				request.session["email_notifications"] = email_notifications
	except requests.RequestException:
		logger.exception("Failed to fetch student profile data")
		api_error = "Could not reach Academic Analyzer API. Please try again later."
	
	# Handle form submission
	if request.method == "POST":
		try:
			form_type = request.POST.get("form_type")
			
			# Handle General Info Form
			if form_type == "general_info":
				new_name = request.POST.get("student_name")
				new_email = request.POST.get("student_email")
				new_email_notifications = "email_notifications" in request.POST
				
				# Validate email
				if not new_email:
					messages.error(request, "Email is required.")
					return redirect("academic_integration:student_profile")
				
				# Update profile through API
				update_data = {
					"rollno": student_roll_number,
					"email": new_email,
					"email_notifications": new_email_notifications
				}
				
				# Only include name if editing is allowed
				if allow_name_edit and new_name:
					update_data["name"] = new_name
				
				response = requests.post(
					f"{api_base_url()}/student/update-profile",
					json=update_data,
					timeout=5,
				)
				
				if response.ok:
					data = _safe_json(response)
					if data.get('success'):
						# Update session data
						if allow_name_edit and new_name:
							request.session["student_name"] = new_name
						request.session["student_email"] = new_email
						request.session["email_notifications"] = new_email_notifications
						
						messages.success(request, "Profile information updated successfully.")
						return redirect("academic_integration:student_profile")
					else:
						messages.error(request, data.get('message', "Failed to update profile."))
				else:
					messages.error(request, "Failed to update profile. Please try again later.")
			
			# Handle Change Password Form
			elif form_type == "change_password":
				new_password = request.POST.get("student_password")
				new_password_confirm = request.POST.get("student_password_confirm")
				
				# Validate passwords
				if not new_password:
					messages.error(request, "Please enter a new password.")
					return redirect("academic_integration:student_profile")
				
				if new_password != new_password_confirm:
					messages.error(request, "Passwords do not match.")
					return redirect("academic_integration:student_profile")
				
				if len(new_password) < 6:
					messages.error(request, "Password must be at least 6 characters long.")
					return redirect("academic_integration:student_profile")
				
				# Update password through API
				update_data = {
					"rollno": student_roll_number,
					"email": student_email,  # Keep existing email
					"password": new_password
				}
				
				response = requests.post(
					f"{api_base_url()}/student/update-profile",
					json=update_data,
					timeout=5,
				)
				
				if response.ok:
					data = _safe_json(response)
					if data.get('success'):
						messages.success(request, "Password changed successfully. Please use your new password for future logins.")
						return redirect("academic_integration:student_profile")
					else:
						messages.error(request, data.get('message', "Failed to change password."))
				else:
					messages.error(request, "Failed to change password. Please try again later.")
			
		except requests.RequestException:
			logger.exception("Failed to update student profile")
			messages.error(request, "Could not reach Academic Analyzer API. Please try again later.")
	
	context = {
		"student_name": student_name,
		"student_roll_number": student_roll_number,
		"student_email": student_email,
		"allow_name_edit": allow_name_edit,
		"email_notifications": email_notifications,
		"api_error": api_error,
	}
	
	return render(request, "academic_integration/student_profile.html", context)


def generate_questions_from_content(request: HttpRequest) -> HttpResponse:
	"""
	API endpoint to generate quiz questions from uploaded content using Gemini API.
	Requires staff authentication.
	"""
	from django.http import JsonResponse
	import json
	import base64
	import io
	import logging
	
	# Set up dedicated logger for this function
	logger = logging.getLogger(__name__)
	
	# Ensure staff is logged in
	if not request.session.get('staff_email'):
		return JsonResponse({'success': False, 'error': 'Not authenticated as staff'}, status=401)
	
	if request.method != 'POST':
		return JsonResponse({'success': False, 'error': 'Only POST method is allowed'}, status=405)
	
	try:
		# Parse request data
		logger.info("Processing question generation request")
		data = json.loads(request.body)
		file_content = data.get('fileContent')
		file_type = data.get('fileType')
		num_questions = int(data.get('numQuestions', 5))
		difficulty = data.get('difficulty', 'medium')
		question_types = data.get('questionTypes', ['mcq_single', 'mcq_multiple', 'true_false'])
		
		# Log parameters (excluding file content)
		logger.info(f"Parameters: file_type={file_type}, num_questions={num_questions}, "
				   f"difficulty={difficulty}, question_types={question_types}")
		
		# Validate required fields
		if not file_content:
			logger.warning("No file content provided in request")
			return JsonResponse({'success': False, 'error': 'No file content provided'}, status=400)
		
		# Check that file type is present
		if not file_type:
			logger.warning("No file type provided in request")
			return JsonResponse({'success': False, 'error': 'No file type provided'}, status=400)
		
		logger.info(f"Received file of type: {file_type}, generating questions...")
		
		# Import our question generator utility
		try:
			from academic_integration.utils.gemini_generator import GeminiQuestionGenerator, extract_text_from_file
		except ImportError as e:
			logger.error(f"Failed to import required modules: {e}")
			return JsonResponse({'success': False, 'error': f'Server configuration error: {str(e)}'}, status=500)
		
		# Use the dedicated file content extraction function from our utility
		try:
			# For direct file upload, we need to use the generator class's file extraction
			generator = GeminiQuestionGenerator()
			
			# Generate questions directly from the file content
			result = generator.generate_questions_from_file(
				file_content=file_content,  # Pass the raw base64 content
				file_type=file_type,
				num_questions=num_questions,
				difficulty=difficulty,
				question_types=question_types
			)
			
			if result.get('success'):
				logger.info(f"Successfully generated {len(result.get('questions', []))} questions")
			else:
				logger.warning(f"Question generation failed: {result.get('error', 'Unknown error')}")
			
			return JsonResponse(result)
			
		except Exception as e:
			logger.exception(f"Error in content extraction or question generation: {e}")
			return JsonResponse({
				'success': False, 
				'error': f'Error processing file: {str(e)}',
				'details': 'Error occurred during content extraction or question generation'
			}, status=400)
	
	except json.JSONDecodeError as e:
		logger.error(f"Invalid JSON in request: {e}")
		return JsonResponse({'success': False, 'error': 'Invalid JSON in request body'}, status=400)
	except Exception as e:
		logger.exception(f"Unexpected error in generate_questions_from_content: {e}")
		return JsonResponse({'success': False, 'error': str(e)}, status=500)


def download_marks_template(request: HttpRequest, course_id: str) -> HttpResponse:
	"""
	Download a CSV template with enrolled students for bulk marks upload.
	Allows selection of which columns to include.
	"""
	import csv
	from django.http import HttpResponse
	
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")
	
	# Get course details and students
	try:
		response = requests.get(
			f"{api_base_url()}/staff/course-detail",
			params={"courseId": course_id},
			timeout=5,
		)
	except requests.RequestException:
		logger.exception("Failed to load course details")
		messages.error(request, "Could not reach Academic Analyzer API.")
		return redirect("academic_integration:manage_course", course_id=course_id)
	
	body = _safe_json(response)
	if not (response.ok and body.get("success")):
		messages.error(request, "Failed to load course details.")
		return redirect("academic_integration:manage_course", course_id=course_id)
	
	students = body.get("students", [])
	
	# Sort students by roll number
	students = sorted(students, key=lambda x: x.get('rollno', ''))
	
	# Get selected columns from query parameters
	include_tutorial1 = request.GET.get('include_tutorial1') == 'on'
	include_tutorial2 = request.GET.get('include_tutorial2') == 'on'
	include_tutorial3 = request.GET.get('include_tutorial3') == 'on'
	include_tutorial4 = request.GET.get('include_tutorial4') == 'on'
	include_ca1 = request.GET.get('include_ca1') == 'on'
	include_ca2 = request.GET.get('include_ca2') == 'on'
	include_assignment = request.GET.get('include_assignment') == 'on'
	
	# Build header based on selections
	header = ['Roll Number', 'Name', 'Email']
	
	# Map of column names
	column_map = []
	if include_tutorial1:
		header.append('Tutorial 1')
		column_map.append('tutorial1')
	if include_tutorial2:
		header.append('Tutorial 2')
		column_map.append('tutorial2')
	if include_tutorial3:
		header.append('Tutorial 3')
		column_map.append('tutorial3')
	if include_tutorial4:
		header.append('Tutorial 4')
		column_map.append('tutorial4')
	if include_ca1:
		header.append('CA 1')
		column_map.append('ca1')
	if include_ca2:
		header.append('CA 2')
		column_map.append('ca2')
	if include_assignment:
		header.append('Assignment/Presentation')
		column_map.append('assignment')
	
	# Create CSV response
	response = HttpResponse(content_type='text/csv')
	response['Content-Disposition'] = f'attachment; filename="marks_template_{course_id}.csv"'
	
	writer = csv.writer(response)
	
	# Write header
	writer.writerow(header)
	
	# Write student rows with empty mark columns
	for student in students:
		row = [
			student.get('rollno', ''),
			student.get('name', ''),
			student.get('email', '')
		]
		# Add empty cells for each selected column
		row.extend(['' for _ in column_map])
		writer.writerow(row)
	
	return response


def download_students_template(request: HttpRequest) -> HttpResponse:
	"""
	Download a CSV template with all students in the system for course enrollment.
	"""
	import csv
	from django.http import HttpResponse
	
	logger.info("Download students template requested")
	
	staff_email = request.session.get("staff_email")
	if not staff_email:
		logger.warning("No staff email in session")
		messages.info(request, "Please log in to continue.")
		return redirect("academic_integration:staff_login")
	
	logger.info(f"Staff email: {staff_email}")
	
	# Get all students from API
	try:
		api_url = f"{api_base_url()}/staff/all-students"
		logger.info(f"Fetching students from: {api_url}")
		response = requests.get(
			api_url,
			params={"email": staff_email},
			timeout=10,
		)
		logger.info(f"API response status: {response.status_code}")
	except requests.RequestException as e:
		logger.exception(f"Failed to load all students: {e}")
		messages.error(request, "Could not reach Academic Analyzer API.")
		return redirect("academic_integration:staff_dashboard")
	
	body = _safe_json(response)
	logger.info(f"API response: {body}")
	
	if not (response.ok and body.get("success")):
		logger.error(f"Failed to load students: {body.get('message')}")
		messages.error(request, "Failed to load students list.")
		return redirect("academic_integration:staff_dashboard")
	
	students = body.get("students", [])
	logger.info(f"Found {len(students)} students")
	
	# Create CSV response
	response = HttpResponse(content_type='text/csv')
	response['Content-Disposition'] = 'attachment; filename="all_students_template.csv"'
	
	writer = csv.writer(response)
	
	# Write header with instruction
	writer.writerow(['Roll Number', 'Name', 'Batch', 'Email'])
	writer.writerow([])  # Empty row
	writer.writerow(['# Keep only the roll numbers you want to add to the course'])
	writer.writerow(['# Delete the Name, Batch, and Email columns before uploading'])
	writer.writerow([])  # Empty row
	
	# Write all students
	for student in students:
		writer.writerow([
			student.get('rollno', ''),
			student.get('name', ''),
			student.get('batch', ''),
			student.get('email', '')
		])
	
	logger.info("CSV template generated successfully")
	return response


# ==================== ARCHIVE MANAGEMENT VIEWS ====================

def archive_course(request: HttpRequest, course_id: str) -> HttpResponse:
	"""Archive a course - move it to archived collection in MongoDB"""
	
	# Check if staff is logged in via session
	staff_email = request.session.get("staff_email")
	logger.info(f"Archive course request - Email from session: {staff_email}")
	logger.info(f"Archive course request - Method: {request.method}")
	
	if not staff_email:
		logger.warning("No staff_email in session, redirecting to login")
		messages.error(request, "You must be logged in as staff")
		return redirect("academic_integration:staff_login")
	
	if request.method != "POST":
		logger.warning(f"Invalid method: {request.method}")
		return HttpResponseBadRequest("Only POST method allowed")
	
	logger.info(f"Attempting to archive course: {course_id} by {staff_email}")
	
	try:
		# Call Academic Analyzer API to archive the course
		response = requests.post(
			f"{api_base_url()}/staff/archive-course",
			json={"email": staff_email, "courseId": course_id},
			timeout=10,
		)
		
		logger.info(f"Archive API response: {response.status_code}")
		
		if response.status_code == 200:
			body = response.json()
			if body.get("success"):
				logger.info(f"Course {course_id} archived successfully")
				messages.success(request, f"Course {course_id} has been archived successfully!")
				return redirect("academic_integration:staff_dashboard")
			else:
				logger.error(f"Archive failed: {body.get('message', 'Unknown error')}")
				messages.error(request, f"Failed to archive course: {body.get('message', 'Unknown error')}")
		else:
			logger.error(f"Archive API error: {response.status_code}")
			messages.error(request, f"API error: {response.status_code}")
	
	except requests.exceptions.RequestException as e:
		logger.error(f"Error archiving course: {e}")
		messages.error(request, "Failed to connect to Academic Analyzer API")
	
	return redirect("academic_integration:manage_course", course_id=course_id)


def restore_course(request: HttpRequest, archived_course_id: str) -> HttpResponse:
	"""Restore an archived course back to active courses"""
	
	# Check if staff is logged in via session
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.error(request, "You must be logged in as staff")
		return redirect("academic_integration:staff_login")
	
	try:
		# Call Academic Analyzer API to restore the course
		response = requests.post(
			f"{api_base_url()}/staff/restore-course",
			json={"email": staff_email, "archivedCourseId": archived_course_id},
			timeout=10,
		)
		
		if response.status_code == 200:
			body = response.json()
			if body.get("success"):
				messages.success(request, f"Course has been restored successfully!")
				return redirect("academic_integration:staff_dashboard")
			else:
				messages.error(request, f"Failed to restore course: {body.get('message', 'Unknown error')}")
		else:
			messages.error(request, f"API error: {response.status_code}")
	
	except requests.exceptions.RequestException as e:
		logger.error(f"Error restoring course: {e}")
		messages.error(request, "Failed to connect to Academic Analyzer API")
	
	return redirect("academic_integration:archived_courses")


def archived_courses(request: HttpRequest) -> HttpResponse:
	"""Display all archived courses for the logged-in staff member"""
	
	# Check if staff is logged in via session
	staff_email = request.session.get("staff_email")
	logger.info(f"Archived courses request - Email from session: {staff_email}")
	
	if not staff_email:
		logger.warning("No staff_email in session for archived courses, redirecting to login")
		messages.error(request, "You must be logged in as staff")
		return redirect("academic_integration:staff_login")
	
	try:
		# Fetch archived courses from API
		response = requests.get(
			f"{api_base_url()}/staff/archived-courses",
			params={"email": staff_email},
			timeout=10,
		)
		
		if response.status_code == 200:
			body = response.json()
			if body.get("success"):
				archived_courses_list = body.get("archivedCourses", [])
			else:
				archived_courses_list = []
				messages.warning(request, "No archived courses found")
		else:
			archived_courses_list = []
			messages.error(request, "Failed to fetch archived courses")
	
	except requests.exceptions.RequestException as e:
		logger.error(f"Error fetching archived courses: {e}")
		archived_courses_list = []
		messages.error(request, "Failed to connect to Academic Analyzer API")
	
	context = {
		"archived_courses": archived_courses_list,
		"staff_name": request.session.get("staff_name", "Staff"),
	}
	
	return render(request, "academic_integration/archived_courses.html", context)


def archived_course_detail(request: HttpRequest, archived_course_id: str) -> HttpResponse:
	"""Display detailed view of an archived course (READ-ONLY)"""
	
	# Check if staff is logged in via session
	staff_email = request.session.get("staff_email")
	if not staff_email:
		messages.error(request, "You must be logged in as staff")
		return redirect("academic_integration:staff_login")
	
	try:
		# Fetch archived course details from API
		response = requests.get(
			f"{api_base_url()}/staff/archived-course-detail",
			params={"archivedCourseId": archived_course_id},
			timeout=10,
		)
		
		if response.status_code == 200:
			body = response.json()
			if body.get("success"):
				course_data = body.get("course")
				
				# Sort students by roll number
				students = course_data.get("students", [])
				students.sort(key=lambda s: s.get("rollno", ""))
				course_data["students"] = students
				
				context = {
					"course": course_data,
					"archived_course_id": archived_course_id,
					"staff_name": request.session.get("staff_name", "Staff"),
				}
				
				return render(request, "academic_integration/archived_course_detail.html", context)
			else:
				messages.error(request, "Archived course not found")
		else:
			messages.error(request, f"API error: {response.status_code}")
	
	except requests.exceptions.RequestException as e:
		logger.error(f"Error fetching archived course detail: {e}")
		messages.error(request, "Failed to connect to Academic Analyzer API")
	
	return redirect("academic_integration:archived_courses")
